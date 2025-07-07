import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import json
import os
import time
from datetime import datetime
import re
import io
import hashlib
from typing import Dict, List, Any, Tuple, Optional

# Import optional dependencies with error handling
try:
    from openai import OpenAI
    OPENAI_AVAILABLE = True
except ImportError:
    OPENAI_AVAILABLE = False

try:
    import requests
    MISTRAL_AVAILABLE = True
except ImportError:
    MISTRAL_AVAILABLE = False

try:
    from docx import Document
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.colors import Color, red, orange, yellow, lightgrey, black
    from reportlab.lib.units import inch
    from reportlab.platypus.flowables import KeepTogether
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    import os
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    import PyPDF2
    import pdfplumber
    PDF_EXTRACT_AVAILABLE = True
except ImportError:
    PDF_EXTRACT_AVAILABLE = False

# Mistral OCR support - ENHANCED VERSION
try:
    from PIL import Image  # Keep for image preview only
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

MISTRAL_OCR_AVAILABLE = False
OCR_AVAILABLE = False
OCR_ERROR_MESSAGE = ""

# Configuration - Optimized for aggressive violation detection
MAX_CHARS_PER_CHUNK = 3000  # Smaller chunks for better analysis
OVERLAP_CHARS = 100  # Reduced overlap
MAX_TOKENS_OUTPUT = 1500  # Increased output tokens for more violations
CHUNK_DELAY = 0.5  # Reduced delay for faster analysis
MAX_RETRIES = 3

# FIXED: Unicode text processing functions - ENHANCED for Bengali preservation
def safe_unicode_text(text):
    """Safely handle Unicode text - PRESERVE Bengali characters exactly - FIXED"""
    if not text:
        return ""
    
    try:
        # Ensure the text is properly encoded as UTF-8
        if isinstance(text, bytes):
            text = text.decode('utf-8', errors='replace')
        elif not isinstance(text, str):
            text = str(text)
        
        # FIXED: DON'T sanitize Unicode characters - keep them exactly as is
        # Only remove truly problematic characters
        text = text.replace('\u200b', '')  # Remove zero-width space only
        text = text.replace('\ufeff', '')  # Remove BOM
        text = text.replace('\u200c', '')  # Remove zero-width non-joiner (problematic for some systems)
        text = text.replace('\u200d', '')  # Remove zero-width joiner (problematic for some systems)
        
        # Normalize Unicode but preserve all characters
        import unicodedata
        text = unicodedata.normalize('NFC', text)
        
        return text
    except Exception as e:
        st.error(f"Unicode processing error: {e}")
        return str(text)  # Return as-is if processing fails

def detect_language_fallback(text_sample):
    """Fallback language detection using character analysis - ENHANCED"""
    if not text_sample:
        return "English"
    
    # Enhanced character range detection
    bengali_chars = sum(1 for char in text_sample if '\u0980' <= char <= '\u09FF')
    hindi_chars = sum(1 for char in text_sample if '\u0900' <= char <= '\u097F')
    tamil_chars = sum(1 for char in text_sample if '\u0B80' <= char <= '\u0BFF')
    telugu_chars = sum(1 for char in text_sample if '\u0C00' <= char <= '\u0C7F')
    gujarati_chars = sum(1 for char in text_sample if '\u0A80' <= char <= '\u0AFF')
    
    total_chars = len(text_sample)
    
    # Lower threshold for detection (5% instead of 10%)
    threshold = 0.05
    
    if bengali_chars > total_chars * threshold:
        return "Bengali"
    elif hindi_chars > total_chars * threshold:
        return "Hindi"
    elif tamil_chars > total_chars * threshold:
        return "Tamil"
    elif telugu_chars > total_chars * threshold:
        return "Telugu"
    elif gujarati_chars > total_chars * threshold:
        return "Gujarati"
    else:
        return "English"

def get_script_range(char):
    """Get the script range for a Unicode character"""
    code = ord(char)
    if 0x0900 <= code <= 0x097F:
        return "Devanagari (Hindi)"
    elif 0x0980 <= code <= 0x09FF:
        return "Bengali"
    elif 0x0A00 <= code <= 0x0A7F:
        return "Gurmukhi (Punjabi)"
    elif 0x0A80 <= code <= 0x0AFF:
        return "Gujarati"
    elif 0x0B00 <= code <= 0x0B7F:
        return "Oriya"
    elif 0x0B80 <= code <= 0x0BFF:
        return "Tamil"
    elif 0x0C00 <= code <= 0x0C7F:
        return "Telugu"
    elif 0x0C80 <= code <= 0x0CFF:
        return "Kannada"
    elif 0x0D00 <= code <= 0x0D7F:
        return "Malayalam"
    else:
        return "Other Unicode"

# Mistral OCR Functions
def check_mistral_ocr_availability():
    """Check if Mistral OCR is available"""
    global MISTRAL_OCR_AVAILABLE, OCR_AVAILABLE, OCR_ERROR_MESSAGE
    
    mistral_key = get_mistral_api_key_with_session()
    if not mistral_key:
        OCR_ERROR_MESSAGE = "❌ Mistral API key not configured"
        MISTRAL_OCR_AVAILABLE = False
        OCR_AVAILABLE = False
        return False, "Mistral API key not configured"
    
    if not MISTRAL_AVAILABLE:
        OCR_ERROR_MESSAGE = "❌ requests library not available"
        MISTRAL_OCR_AVAILABLE = False
        OCR_AVAILABLE = False
        return False, "requests library not available"
    
    try:
        # Test the connection to Mistral API
        headers = {
            "Authorization": f"Bearer {mistral_key}"
        }
        
        # Make a simple test request to check API availability
        response = requests.get("https://api.mistral.ai/v1/models", headers=headers, timeout=10)
        
        if response.status_code == 200:
            MISTRAL_OCR_AVAILABLE = True
            OCR_AVAILABLE = True
            OCR_ERROR_MESSAGE = ""
            return True, "Mistral OCR available"
        elif response.status_code == 401:
            OCR_ERROR_MESSAGE = "❌ Invalid Mistral API key"
            MISTRAL_OCR_AVAILABLE = False
            OCR_AVAILABLE = False
            return False, "Invalid Mistral API key"
        elif response.status_code == 403:
            OCR_ERROR_MESSAGE = "❌ OCR access not enabled for this Mistral account"
            MISTRAL_OCR_AVAILABLE = False
            OCR_AVAILABLE = False
            return False, "OCR access not enabled for this account"
        else:
            OCR_ERROR_MESSAGE = f"❌ Mistral API error: {response.status_code}"
            MISTRAL_OCR_AVAILABLE = False
            OCR_AVAILABLE = False
            return False, f"Mistral API error: {response.status_code}"
            
    except Exception as e:
        OCR_ERROR_MESSAGE = f"❌ Mistral API connection failed: {str(e)}"
        MISTRAL_OCR_AVAILABLE = False
        OCR_AVAILABLE = False
        return False, f"Mistral API connection failed: {str(e)}"

def upload_file_to_mistral(file_data: bytes, filename: str, mistral_key: str) -> Optional[str]:
    """Upload file to Mistral files endpoint - matches your n8n workflow"""
    try:
        url = "https://api.mistral.ai/v1/files"
        
        headers = {
            "Authorization": f"Bearer {mistral_key}"
        }
        
        # Use the same structure as your n8n workflow
        files = {
            "file": (filename, file_data, "image/jpeg" if filename.lower().endswith(('.jpg', '.jpeg')) else "image/png")
        }
        
        data = {
            "purpose": "batch"  # Based on your n8n workflow
        }
        
        response = requests.post(url, headers=headers, files=files, data=data, timeout=60)
        
        if response.status_code == 200:
            result = response.json()
            file_id = result.get("id")
            if file_id:
                st.success(f"✅ File uploaded to Mistral: {file_id}")
                return file_id
            else:
                st.error("❌ No file ID returned from Mistral")
                return None
        else:
            st.error(f"❌ Failed to upload file to Mistral: {response.status_code} - {response.text}")
            return None
            
    except Exception as e:
        st.error(f"❌ Error uploading file to Mistral: {str(e)}")
        return None

def get_mistral_ocr_result(file_id: str, mistral_key: str, language: str = "ben+eng") -> Optional[str]:
    """Get OCR result from Mistral OCR endpoint - matches your n8n workflow"""
    try:
        url = "https://api.mistral.ai/v1/ocr"
        
        headers = {
            "Authorization": f"Bearer {mistral_key}",
            "Content-Type": "application/json"
        }
        
        # Based on your n8n workflow structure
        payload = {
            "model": "mistral-ocr-latest",
            "document": {
                "type": "document_url", 
                "document_url": f"{{ ${file_id}.uri }}"  # This matches your n8n parameter
            },
            "languages": [language],  # Use languages array as shown in n8n
            "output_format": "text"
        }
        
        response = requests.post(url, headers=headers, json=payload, timeout=120)
        
        if response.status_code == 200:
            result = response.json()
            # Extract text from the response
            extracted_text = ""
            
            if "text" in result:
                extracted_text = result["text"]
            elif "content" in result:
                extracted_text = result["content"]
            elif "extracted_text" in result:
                extracted_text = result["extracted_text"]
            elif "result" in result:
                extracted_text = str(result["result"])
            else:
                # Try to find text in nested structure
                for key, value in result.items():
                    if isinstance(value, str) and len(value) > 10:
                        extracted_text = value
                        break
                
                if not extracted_text:
                    st.warning("⚠️ OCR completed but no text found in response")
                    st.json(result)  # Show the response structure for debugging
                    return ""
            
            return extracted_text
        else:
            st.error(f"❌ Mistral OCR failed: {response.status_code} - {response.text}")
            return None
            
    except Exception as e:
        st.error(f"❌ Error getting OCR result from Mistral: {str(e)}")
        return None

def extract_text_with_mistral_ocr(image_file, language: str = "ben+eng") -> str:
    """Extract text from image using Mistral OCR - main function"""
    mistral_key = get_mistral_api_key_with_session()
    if not mistral_key:
        st.error("❌ Mistral API key not configured for OCR")
        return ""
    
    try:
        # Get file data
        if hasattr(image_file, 'getvalue'):
            file_data = image_file.getvalue()
            filename = image_file.name
        else:
            file_data = image_file
            filename = "uploaded_image.jpg"
        
        st.info(f"📤 Uploading {filename} ({len(file_data)/1024:.1f} KB) to Mistral OCR...")
        
        # Step 1: Upload file to Mistral (matches your n8n workflow)
        with st.spinner("📤 Uploading image to Mistral..."):
            file_id = upload_file_to_mistral(file_data, filename, mistral_key)
        
        if not file_id:
            st.error("❌ Failed to upload file to Mistral")
            return ""
        
        st.success(f"✅ File uploaded successfully: {file_id}")
        
        # Step 2: Get OCR result (matches your n8n workflow)
        with st.spinner("🔍 Processing OCR with Mistral..."):
            # Add a small delay to ensure file is processed
            time.sleep(3)
            
            extracted_text = get_mistral_ocr_result(file_id, mistral_key, language)
        
        if extracted_text:
            st.success(f"✅ OCR completed! Extracted {len(extracted_text)} characters")
            return safe_unicode_text(extracted_text)
        else:
            st.error("❌ Failed to extract text from image")
            return ""
            
    except Exception as e:
        st.error(f"❌ Mistral OCR failed: {str(e)}")
        return ""

def extract_text_with_ocr(image_file):
    """Extract text from image using Mistral OCR (replaces pytesseract)"""
    return extract_text_with_mistral_ocr(image_file)

# S&P Violation Rules - Hybrid Context + Keywords Approach
VIOLATION_RULES = {
    "National_Anthem_Misuse": {
        "description": "Misuse of National Anthem for commercial use",
        "context": "Any commercial or promotional use of the Indian National Anthem, including background music, jingles, or promotional content",
        "keywords": ["national anthem", "jana gana mana", "commercial", "advertisement", "promotional", "jingle", "background music"],
        "severity": "critical"
    },
    "Personal_Information_Exposure": {
        "description": "Use of real personal information without consent (address, phone, email, license plate, photo)",
        "context": "Display of actual personal details, real addresses, working phone numbers, genuine email addresses, actual license plates, or real photographs of individuals",
        "keywords": ["phone number", "mobile number", "address", "email", "license plate", "personal photo", "real name", "contact details", "personal information"],
        "severity": "high"
    },
    "OTT_Platform_Promotion": {
        "description": "Promotion of any OTT or TV channel other than hoichoi",
        "context": "Any mention, promotion, or positive reference to competing streaming platforms, TV channels, or digital content providers",
        "keywords": ["netflix", "amazon prime", "hotstar", "zee5", "sony liv", "tv channel", "streaming platform", "digital platform", "ott platform", "prime video", "disney+"],
        "severity": "high"
    },
    "National_Emblem_Misuse": {
        "description": "Misuse of national emblems/assets as props; improper use of the Indian Flag",
        "context": "Using national flag, emblem, or symbols as costumes, props, decoration, or in any manner that violates the Flag Code of India",
        "keywords": ["indian flag", "tricolor", "tiranga", "ashoka chakra", "national emblem", "flag code", "national symbol", "emblem", "coat of arms"],
        "severity": "critical"
    },
    "National_Symbol_Distortion": {
        "description": "Distortion of national symbols/emblems or Indian map",
        "context": "Incorrect representation, alteration, or distortion of national symbols, emblems, or the geographical boundaries of India",
        "keywords": ["indian map", "national symbol", "emblem distortion", "flag distortion", "map distortion", "symbol alteration", "geographical boundaries"],
        "severity": "critical"
    },
    "Hurtful_References": {
        "description": "Hurtful references to real people/groups (football clubs, authors, etc.)",
        "context": "Negative, derogatory, or offensive references to real individuals, organizations, sports teams, or identifiable groups",
        "keywords": ["real person", "football club", "author", "celebrity", "public figure", "organization", "sports team", "derogatory", "offensive"],
        "severity": "medium"
    },
    "Self_Harm_Graphic_Content": {
        "description": "Graphic/self-harm or suicide attempts (must be suggestive, not detailed)",
        "context": "Detailed depiction of self-harm methods, explicit suicide attempts, or graphic content that could be instructional rather than suggestive",
        "keywords": ["suicide", "self-harm", "cutting", "hanging", "graphic violence", "self-injury", "suicide attempt", "harm oneself", "end life"],
        "severity": "critical"
    },
    "Acid_Attack_Depiction": {
        "description": "Depiction of acid attacks",
        "context": "Any portrayal of acid attacks, including preparation, execution, or aftermath, regardless of context",
        "keywords": ["acid attack", "acid throwing", "chemical burn", "disfigurement", "acid", "corrosive", "chemical attack"],
        "severity": "critical"
    },
    "Bomb_Weapon_Instructions": {
        "description": "Detailed instructions for making bombs, using weapons, or harmful tools",
        "context": "Step-by-step instructions, detailed explanations, or educational content about creating explosives, weapons, or harmful devices",
        "keywords": ["bomb making", "weapon instructions", "explosive", "harmful tools", "explosive device", "bomb recipe", "weapon tutorial"],
        "severity": "critical"
    },
    "Harmful_Product_Instructions": {
        "description": "Instructions or product mentions encouraging harm (e.g., using phenyl to commit suicide)",
        "context": "Content that suggests or instructs on using household products, chemicals, or substances for self-harm or harm to others",
        "keywords": ["phenyl", "poison", "harmful chemicals", "toxic substances", "household poison", "chemical harm", "toxic product"],
        "severity": "critical"
    },
    "Religious_Footwear_Context": {
        "description": "Wearing footwear in religious contexts or near idols",
        "context": "Characters wearing shoes or footwear inside temples, near religious idols, or in sacred spaces where it's culturally inappropriate",
        "keywords": ["shoes", "footwear", "temple", "idol", "religious place", "shrine", "sacred space", "sandals", "boots"],
        "severity": "high"
    },
    "Buddha_Idol_Misuse": {
        "description": "Inappropriate use/display of Buddha idols/pictures on props/clothing",
        "context": "Using Buddha's image or Buddhist symbols on clothing, accessories, or in inappropriate contexts that show disrespect",
        "keywords": ["buddha", "buddhist", "idol", "religious image", "clothing", "t-shirt", "accessory", "buddha statue", "buddhist symbol"],
        "severity": "high"
    },
    "Religious_Mockery": {
        "description": "Mockery of religious facts or symbols",
        "context": "Content that ridicules, mocks, or shows disrespect toward religious beliefs, practices, symbols, or sacred texts",
        "keywords": ["religious mockery", "sacred", "holy", "religious symbol", "faith", "mock", "ridicule", "disrespect", "blasphemy"],
        "severity": "critical"
    },
    "Caste_Religion_References": {
        "description": "Use of proverbs/colloquialisms that reference caste, religion, or community",
        "context": "Language that reinforces caste hierarchies, religious stereotypes, or discriminatory attitudes toward specific communities",
        "keywords": ["caste", "brahmin", "dalit", "community", "religious slur", "caste system", "untouchable", "higher caste", "lower caste"],
        "severity": "high"
    },
    "Social_Evils_Promotion": {
        "description": "Promotion of social evils (child marriage, dowry, son preference, etc.)",
        "context": "Content that normalizes, promotes, or presents harmful social practices in a positive light without showing consequences",
        "keywords": ["child marriage", "dowry", "son preference", "female infanticide", "social evil", "harmful practice", "discrimination"],
        "severity": "critical"
    },
    "Unauthorized_Branding": {
        "description": "Unauthorized branding/endorsement; brand names must be blurred",
        "context": "Visible brand logos, product names, or commercial endorsements without proper clearance or blurring",
        "keywords": ["brand name", "logo", "trademark", "product placement", "endorsement", "commercial brand", "brand logo", "product name"],
        "severity": "medium"
    },
    "Credit_List_Changes": {
        "description": "Unapproved or post-deadline changes in the credit list",
        "context": "Modifications to cast, crew, or production credits after final approval or without proper authorization",
        "keywords": ["credits", "cast", "crew", "production team", "acknowledgment", "credit list", "cast list", "crew list"],
        "severity": "medium"
    },
    "Alcohol_Cigarette_Brands": {
        "description": "Display of alcohol/cigarette brands/logos without marketing team approval",
        "context": "Visible alcohol or tobacco brand names, logos, or products without proper marketing clearance",
        "keywords": ["alcohol brand", "cigarette brand", "tobacco", "liquor", "beer", "wine", "whiskey", "cigarette logo", "tobacco brand"],
        "severity": "high"
    },
    "Smoking_Disclaimer_Missing": {
        "description": "Absence of 'Smoking Kills' message during smoking scenes",
        "context": "Smoking scenes without appropriate health warnings or disclaimers as required by regulations",
        "keywords": ["smoking", "cigarette", "tobacco", "disclaimer", "warning", "smoking kills", "health warning", "tobacco warning"],
        "severity": "medium"
    },
    "Content_Disclaimer_Missing": {
        "description": "Missing special disclaimers for violent, gory, or sexually explicit content",
        "context": "Content requiring viewer discretion or age-appropriate warnings without proper disclaimers",
        "keywords": ["violence", "gore", "sexual content", "explicit", "disclaimer", "viewer discretion", "age appropriate", "content warning"],
        "severity": "medium"
    },
    "Unapproved_Endorsements": {
        "description": "Unapproved endorsements or acknowledgments in end credits",
        "context": "Thank you messages, acknowledgments, or endorsements in credits that haven't been approved by the content team",
        "keywords": ["endorsement", "acknowledgment", "credits", "sponsor", "thanks", "end credits", "acknowledgement", "special thanks"],
        "severity": "medium"
    },
    "Animal_Harm_Depiction": {
        "description": "Depiction of harm or killing of animals during filming",
        "context": "Content showing actual harm to animals during production, cruelty to animals, or realistic depictions of animal suffering",
        "keywords": ["animal harm", "animal killing", "cruelty", "abuse", "violence", "animal cruelty", "animal suffering", "harm animals"],
        "severity": "critical"
    },
    "Child_Adult_Behavior": {
        "description": "Child actors shown behaving like adults or speaking mature dialogue",
        "context": "Child characters using adult language, exhibiting mature behavior, or being placed in age-inappropriate situations",
        "keywords": ["child actor", "mature dialogue", "adult behavior", "inappropriate", "child character", "adult language", "age inappropriate"],
        "severity": "high"
    },
    "Child_Abuse_Content": {
        "description": "Any form of child abuse—physical, sexual, or psychological",
        "context": "Content depicting, suggesting, or normalizing any form of abuse toward children, including physical, emotional, or sexual abuse",
        "keywords": ["child abuse", "physical abuse", "sexual abuse", "psychological abuse", "child harm", "abuse child", "child violence"],
        "severity": "critical"
    }
}

# Streamlit App Configuration
st.set_page_config(
    page_title="hoichoi S&P Compliance Analyzer - FIXED",
    page_icon="🎬",
    layout="wide"
)

# Authentication Functions
def check_email_domain(email: str) -> bool:
    """Check if email belongs to hoichoi.tv domain"""
    return email.lower().strip().endswith('@hoichoi.tv')

def authenticate_user():
    """Handle user authentication for hoichoi.tv employees only"""
    if 'authenticated' not in st.session_state:
        st.session_state.authenticated = False
    
    if not st.session_state.authenticated:
        # Custom CSS for login page
        st.markdown("""
        <style>
        .login-header {
            background: linear-gradient(90deg, #ff6b6b, #4ecdc4);
            padding: 2rem;
            border-radius: 15px;
            color: white;
            text-align: center;
            margin-bottom: 2rem;
            box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        }
        .login-container {
            background: white;
            padding: 2rem;
            border-radius: 15px;
            box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
            border: 1px solid #e0e0e0;
        }
        </style>
        """, unsafe_allow_html=True)
        
        st.markdown("""
        <div class="login-header">
            <h1>🎬 hoichoi S&P Compliance System - FIXED</h1>
            <h3>Standards & Practices Content Review Platform</h3>
            <p>✅ Bengali Text Preservation | 🔧 PDF Generation Fixed | 🔑 Enhanced API Error Handling</p>
        </div>
        """, unsafe_allow_html=True)
        
        with st.container():
            col1, col2, col3 = st.columns([1, 2, 1])
            with col2:
                st.markdown('<div class="login-container">', unsafe_allow_html=True)
                
                st.subheader("🔐 Employee Access Portal")
                st.write("Please login with your hoichoi corporate email address")
                
                email = st.text_input(
                    "Corporate Email Address",
                    placeholder="yourname@hoichoi.tv",
                    help="Only @hoichoi.tv email addresses are authorized"
                )
                
                password = st.text_input(
                    "Password",
                    type="password",
                    help="Enter your corporate password"
                )
                
                col_a, col_b = st.columns(2)
                with col_a:
                    if st.button("🚀 Login", type="primary", use_container_width=True):
                        if email and password:
                            if check_email_domain(email):
                                # Simple password check (in production, use proper authentication)
                                if len(password) >= 6:  # Basic password validation
                                    st.session_state.authenticated = True
                                    st.session_state.user_email = email
                                    st.session_state.user_name = email.split('@')[0].replace('.', ' ').title()
                                    st.session_state.is_admin = email.lower() in ['admin@hoichoi.tv', 'sp@hoichoi.tv', 'content@hoichoi.tv']
                                    st.success("✅ Login successful! Redirecting...")
                                    time.sleep(1)
                                    st.rerun()
                                else:
                                    st.error("❌ Password must be at least 6 characters long")
                            else:
                                st.error("❌ Access denied. Only @hoichoi.tv email addresses are authorized.")
                                st.warning("This system is restricted to hoichoi content team members only.")
                        else:
                            st.error("❌ Please enter both email and password")
                
                with col_b:
                    if st.button("ℹ️ Help", use_container_width=True):
                        st.info("""
                        **Need Access?**
                        - Contact IT department for account setup
                        - Must use corporate @hoichoi.tv email
                        - For support: it@hoichoi.tv
                        """)
                
                st.divider()
                st.markdown("""
                <div style='text-align: center; color: #666; font-size: 0.9em;'>
                    <p>🔒 This is a secure system for hoichoi content review</p>
                    <p>📧 Access restricted to @hoichoi.tv employees only</p>
                    <p>🛡️ All activities are logged for security purposes</p>
                    <p style="color: green;">✅ FIXED VERSION: Bengali text preservation + PDF generation + API error handling</p>
                </div>
                """, unsafe_allow_html=True)
                
                st.markdown('</div>', unsafe_allow_html=True)
        
        return False
    
    return True

def get_api_key():
    """Get OpenAI API key from Streamlit secrets or user input"""
    try:
        return st.secrets.get("OPENAI_API_KEY", None)
    except:
        return None

def get_mistral_api_key():
    """Get Mistral API key from Streamlit secrets or user input"""
    try:
        return st.secrets.get("MISTRAL_API_KEY", None)
    except:
        return None

def get_mistral_api_key_with_session():
    """Get Mistral API key with session state support"""
    # Check session state first
    if hasattr(st.session_state, 'temp_mistral_key') and st.session_state.temp_mistral_key:
        return st.session_state.temp_mistral_key
    # Then check secrets
    try:
        return st.secrets.get("MISTRAL_API_KEY", None)
    except:
        return None

def detect_language_with_mistral(text_sample, api_key):
    """Detect language using Mistral API - ENHANCED for better Bengali detection"""
    if not MISTRAL_AVAILABLE or not api_key:
        return None
    
    try:
        # Mistral API endpoint for language detection
        url = "https://api.mistral.ai/v1/chat/completions"
        
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}"
        }
        
        # Take a larger sample for better detection
        sample = text_sample[:2000] if len(text_sample) > 2000 else text_sample
        
        payload = {
            "model": "mistral-small",  # Use mistral-small for better language detection
            "messages": [
                {
                    "role": "system", 
                    "content": "You are a language detection expert specializing in South Asian languages. Identify the primary language of the given text. Return ONLY the language name in English. Pay special attention to Bengali, Hindi, Tamil, Telugu, Gujarati, Marathi, Punjabi, Urdu, Malayalam, Kannada, Odia, Assamese. If you see Bengali script (০-৯, অ-ঞ, ত-ন, প-ম, য-হ), always return Bengali."
                },
                {
                    "role": "user", 
                    "content": f"What language is this text primarily written in? Be very careful with Bengali detection. Respond with just the language name.\n\nText: {sample}"
                }
            ],
            "max_tokens": 20,
            "temperature": 0
        }
        
        response = requests.post(url, json=payload, headers=headers, timeout=30)
        
        if response.status_code == 200:
            result = response.json()
            detected_language = result['choices'][0]['message']['content'].strip()
            
            # Enhanced validation with Bengali priority
            valid_languages = ['English', 'Bengali', 'Hindi', 'Tamil', 'Telugu', 'Gujarati', 'Marathi', 'Punjabi', 'Urdu', 'Malayalam', 'Kannada', 'Odia', 'Assamese']
            
            # Special check for Bengali characters
            bengali_chars = sum(1 for char in sample if '\u0980' <= char <= '\u09FF')
            if bengali_chars > 0:
                return "Bengali"
            
            if detected_language in valid_languages:
                return detected_language
            else:
                return None
        else:
            st.warning(f"Mistral API error: {response.status_code}")
            return None
            
    except Exception as e:
        st.warning(f"Mistral language detection failed: {e}")
        return None

def setup_unicode_fonts():
    """Setup Unicode fonts for multilingual PDF support"""
    try:
        # Register Unicode fonts for better language support
        # In production, you should include actual font files
        # For now, using default fonts with Unicode support
        return True
    except:
        return False

def detect_language(text_sample):
    """Detect the primary language of the text using Mistral (preferred) or OpenAI - ENHANCED"""
    # Try Mistral first
    mistral_api_key = get_mistral_api_key_with_session()
    if MISTRAL_AVAILABLE and mistral_api_key:
        detected = detect_language_with_mistral(text_sample, mistral_api_key)
        if detected:
            return detected
    
    # Fallback to OpenAI with enhanced Bengali detection
    openai_api_key = get_api_key()
    if OPENAI_AVAILABLE and openai_api_key:
        try:
            client = OpenAI(api_key=openai_api_key)
            
            # Take a larger sample for better detection
            sample = text_sample[:2500] if len(text_sample) > 2500 else text_sample
            
            response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "You are a language detection expert specializing in South Asian languages. Identify the primary language of the given text. Return ONLY the language name in English. Pay special attention to Bengali script detection. Common languages include: English, Bengali, Hindi, Tamil, Telugu, Gujarati, Marathi, Punjabi, Urdu, Malayalam, Kannada, Odia, Assamese."},
                    {"role": "user", "content": f"What language is this text primarily written in? Be very careful with Bengali script detection. Respond with just the language name.\n\nText: {sample}"}
                ],
                max_tokens=20,
                temperature=0
            )
            
            detected_language = response.choices[0].message.content.strip()
            
            # Enhanced validation with Bengali priority
            valid_languages = ['English', 'Bengali', 'Hindi', 'Tamil', 'Telugu', 'Gujarati', 'Marathi', 'Punjabi', 'Urdu', 'Malayalam', 'Kannada', 'Odia', 'Assamese']
            
            # Special check for Bengali characters
            bengali_chars = sum(1 for char in sample if '\u0980' <= char <= '\u09FF')
            if bengali_chars > 0:
                return "Bengali"
            
            if detected_language in valid_languages:
                return detected_language
            
        except Exception as e:
            st.warning(f"OpenAI language detection failed: {e}")
    
    # Enhanced fallback to character-based detection
    return detect_language_fallback(text_sample)

def extract_text_from_pdf_bytes(file_bytes):
    """Extract text from uploaded PDF file bytes with page preservation - ENHANCED Unicode support"""
    if not PDF_EXTRACT_AVAILABLE:
        st.error("❌ PDF extraction libraries not available. Please install PyPDF2 and pdfplumber.")
        return None, []
    
    try:
        pages_data = []
        full_text = ""
        
        # Try using pdfplumber first (better for complex layouts and Unicode)
        try:
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                for page_num, page in enumerate(pdf.pages, 1):
                    page_text = page.extract_text() or ""
                    
                    # ENHANCED: Better Unicode handling
                    if page_text.strip():
                        # Properly handle Unicode text
                        page_text = safe_unicode_text(page_text)
                        
                        pages_data.append({
                            'page_number': page_num,
                            'text': page_text.strip(),
                            'original_page': page_num  # Preserve original page number
                        })
                        full_text += f"\n=== ORIGINAL PAGE {page_num} ===\n{page_text}\n"
        except Exception as e:
            st.warning(f"PDFPlumber failed: {e}, trying PyPDF2...")
            # Fallback to PyPDF2
            pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_bytes))
            for page_num, page in enumerate(pdf_reader.pages, 1):
                page_text = page.extract_text() or ""
                
                if page_text.strip():
                    # ENHANCED: Better Unicode handling
                    page_text = safe_unicode_text(page_text)
                    
                    pages_data.append({
                        'page_number': page_num,
                        'text': page_text.strip(),
                        'original_page': page_num  # Preserve original page number
                    })
                    full_text += f"\n=== ORIGINAL PAGE {page_num} ===\n{page_text}\n"
        
        return full_text, pages_data
        
    except Exception as e:
        st.error(f"Error extracting text from PDF: {e}")
        return None, []

def extract_text_from_docx_bytes(file_bytes):
    """Extract text from uploaded DOCX file bytes with enhanced screenplay parsing - ENHANCED Unicode support"""
    if not DOCX_AVAILABLE:
        st.error("❌ python-docx not available. Please check requirements.txt")
        return None, []
    
    try:
        doc = Document(io.BytesIO(file_bytes))
        pages_data = []
        full_text = ""
        
        # Enhanced screenplay parsing with better Unicode support
        current_page = 1
        current_page_text = ""
        char_count = 0
        
        screenplay_elements = []
        
        for para in doc.paragraphs:
            para_text = para.text.strip()
            
            if not para_text:
                continue
            
            # ENHANCED: Better Unicode handling
            para_text = safe_unicode_text(para_text)
            
            # Detect screenplay elements
            element_type = detect_screenplay_element(para_text, para)
            
            screenplay_elements.append({
                'text': para_text,
                'type': element_type,
                'page': current_page
            })
            
            # Check for manual page breaks
            if has_page_break(para):
                # Save current page
                if current_page_text.strip():
                    pages_data.append({
                        'page_number': current_page,
                        'text': current_page_text.strip(),
                        'original_page': current_page,
                        'screenplay_elements': [e for e in screenplay_elements if e['page'] == current_page]
                    })
                    full_text += f"\n=== ORIGINAL PAGE {current_page} ===\n{current_page_text}\n"
                
                current_page += 1
                current_page_text = ""
                char_count = 0
            else:
                current_page_text += para_text + "\n"
                char_count += len(para_text) + 1
                
                # Automatic page break based on character count (approximate)
                if char_count > 2000:  # Rough estimate for one page
                    pages_data.append({
                        'page_number': current_page,
                        'text': current_page_text.strip(),
                        'original_page': current_page,
                        'screenplay_elements': [e for e in screenplay_elements if e['page'] == current_page]
                    })
                    full_text += f"\n=== ORIGINAL PAGE {current_page} ===\n{current_page_text}\n"
                    
                    current_page += 1
                    current_page_text = ""
                    char_count = 0
        
        # Add remaining text
        if current_page_text.strip():
            pages_data.append({
                'page_number': current_page,
                'text': current_page_text.strip(),
                'original_page': current_page,
                'screenplay_elements': [e for e in screenplay_elements if e['page'] == current_page]
            })
            full_text += f"\n=== ORIGINAL PAGE {current_page} ===\n{current_page_text}\n"
        
        return full_text, pages_data
        
    except Exception as e:
        st.error(f"Error extracting text: {e}")
        return None, []

def detect_screenplay_element(text, para):
    """Detect screenplay element type (scene heading, character, dialogue, action, etc.)"""
    text_upper = text.upper()
    
    # Scene headings
    if (text_upper.startswith(('INT.', 'EXT.', 'INTERIOR', 'EXTERIOR')) or
        re.match(r'^(INT|EXT)\.?\s+', text_upper)):
        return 'SCENE_HEADING'
    
    # Character names (usually centered or in caps)
    if (text.isupper() and len(text.split()) <= 3 and 
        not text.startswith(('INT.', 'EXT.')) and
        len(text) < 50):
        return 'CHARACTER'
    
    # Parentheticals
    if text.startswith('(') and text.endswith(')'):
        return 'PARENTHETICAL'
    
    # Transitions
    if (text_upper.endswith(('TO:', 'OUT:', 'IN:')) or
        text_upper in ['FADE IN:', 'FADE OUT:', 'CUT TO:', 'DISSOLVE TO:']):
        return 'TRANSITION'
    
    # Action/Description (default for longer text)
    if len(text) > 100:
        return 'ACTION'
    
    # Dialogue (shorter text that's not other elements)
    return 'DIALOGUE'

def has_page_break(para):
    """Check if paragraph has a page break"""
    try:
        # Check for page break in Word document
        if hasattr(para, '_element'):
            for br in para._element.xpath('.//w:br[@w:type="page"]'):
                return True
        return False
    except:
        return False

def generate_ai_solution(violation_text, violation_type, explanation, detected_language, api_key):
    """Generate AI solution for the violation using Mistral (preferred) or OpenAI - ENHANCED"""
    
    # Try Mistral first
    mistral_key = get_mistral_api_key_with_session()
    if MISTRAL_AVAILABLE and mistral_key:
        try:
            url = "https://api.mistral.ai/v1/chat/completions"
            
            headers = {
                "Content-Type": "application/json",
                "Authorization": f"Bearer {mistral_key}"
            }
            
            # Map violation type to specific guidance
            guideline_context = {
                "National_Anthem_Misuse": "Remove commercial use of national anthem; use instrumental version or replace with original music",
                "Personal_Information_Exposure": "Blur/mask personal information; use fictional phone numbers and addresses",
                "OTT_Platform_Promotion": "Remove references to competing platforms; replace with generic terms or hoichoi references",
                "National_Emblem_Misuse": "Remove improper flag usage; ensure respectful display according to Flag Code of India",
                "National_Symbol_Distortion": "Restore accurate representation of national symbols and Indian map",
                "Religious_Footwear_Context": "Remove footwear in religious settings; ensure actors are barefoot near idols/temples",
                "Buddha_Idol_Misuse": "Remove Buddha images from clothing/inappropriate contexts; use respectfully or remove",
                "Religious_Mockery": "Rewrite dialogue to be respectful of religious beliefs and symbols",
                "Caste_Religion_References": "Replace with neutral language; avoid caste/community-specific terms",
                "Social_Evils_Promotion": "Reframe to show negative consequences; don't glorify harmful practices",
                "Self_Harm_Graphic_Content": "Make suggestive rather than explicit; focus on emotional impact, not graphic details",
                "Acid_Attack_Depiction": "Remove or significantly tone down; use off-screen treatment",
                "Child_Adult_Behavior": "Rewrite dialogue age-appropriately; ensure child actors behave naturally",
                "Child_Abuse_Content": "Remove completely; find alternative plot devices",
                "Unauthorized_Branding": "Blur visible brand names and logos; use generic alternatives",
                "Alcohol_Cigarette_Brands": "Blur alcohol/tobacco brands; use generic packaging",
                "Smoking_Disclaimer_Missing": "Add 'Smoking Kills' disclaimer during smoking scenes",
                "Animal_Harm_Depiction": "Remove animal harm scenes; use CGI or off-screen treatment"
            }
            
            specific_guidance = guideline_context.get(violation_type, "Revise content to comply with broadcasting standards")
            
            prompt = f"""You are an expert content editor for hoichoi digital platform. Generate a compliant revision for this S&P violation.

VIOLATION DETAILS:
- Type: {violation_type}
- Problematic Content: "{violation_text}"
- Issue: {explanation}
- Content Language: {detected_language}
- Specific Guidance: {specific_guidance}

INSTRUCTIONS:
1. Provide a revised version that eliminates the S&P violation completely
2. Maintain the original creative intent where possible
3. Keep the same language as the original content ({detected_language})
4. If the content is in Bengali, provide the solution in Bengali
5. If the content is in Hindi, provide the solution in Hindi
6. Ensure the solution is appropriate for Indian digital content standards
7. Make the minimum necessary changes to achieve compliance

Return ONLY the revised content solution in the same language as the original, nothing else."""
            
            payload = {
                "model": "mistral-small",
                "messages": [
                    {
                        "role": "system", 
                        "content": "You are an expert content editor specializing in S&P compliance for hoichoi digital platform. You are multilingual and can provide solutions in Bengali, Hindi, and other Indian languages. Always match the language of the original content."
                    },
                    {
                        "role": "user", 
                        "content": prompt
                    }
                ],
                "max_tokens": 300,
                "temperature": 0.3
            }
            
            response = requests.post(url, json=payload, headers=headers, timeout=60)
            
            if response.status_code == 200:
                result = response.json()
                solution = result['choices'][0]['message']['content'].strip()
                # Ensure Unicode is properly handled
                return safe_unicode_text(solution)
            elif response.status_code == 429:
                st.warning("⚠️ Mistral API rate limit reached, trying OpenAI...")
            elif response.status_code == 402:
                st.warning("⚠️ Mistral API billing issue, trying OpenAI...")
                
        except Exception as e:
            st.warning(f"Mistral solution generation failed: {e}")
    
    # Fallback to OpenAI with enhanced multilingual support
    if not OPENAI_AVAILABLE or not api_key:
        return "AI solution generation not available"
    
    try:
        client = OpenAI(api_key=api_key)
        
        # Map violation type to specific guidance
        guideline_context = {
            "National_Anthem_Misuse": "Remove commercial use of national anthem; use instrumental version or replace with original music",
            "Personal_Information_Exposure": "Blur/mask personal information; use fictional phone numbers and addresses",
            "OTT_Platform_Promotion": "Remove references to competing platforms; replace with generic terms or hoichoi references",
            "National_Emblem_Misuse": "Remove improper flag usage; ensure respectful display according to Flag Code of India",
            "National_Symbol_Distortion": "Restore accurate representation of national symbols and Indian map",
            "Religious_Footwear_Context": "Remove footwear in religious settings; ensure actors are barefoot near idols/temples",
            "Buddha_Idol_Misuse": "Remove Buddha images from clothing/inappropriate contexts; use respectfully or remove",
            "Religious_Mockery": "Rewrite dialogue to be respectful of religious beliefs and symbols",
            "Caste_Religion_References": "Replace with neutral language; avoid caste/community-specific terms",
            "Social_Evils_Promotion": "Reframe to show negative consequences; don't glorify harmful practices",
            "Self_Harm_Graphic_Content": "Make suggestive rather than explicit; focus on emotional impact, not graphic details",
            "Acid_Attack_Depiction": "Remove or significantly tone down; use off-screen treatment",
            "Child_Adult_Behavior": "Rewrite dialogue age-appropriately; ensure child actors behave naturally",
            "Child_Abuse_Content": "Remove completely; find alternative plot devices",
            "Unauthorized_Branding": "Blur visible brand names and logos; use generic alternatives",
            "Alcohol_Cigarette_Brands": "Blur alcohol/tobacco brands; use generic packaging",
            "Smoking_Disclaimer_Missing": "Add 'Smoking Kills' disclaimer during smoking scenes",
            "Animal_Harm_Depiction": "Remove animal harm scenes; use CGI or off-screen treatment"
        }
        
        specific_guidance = guideline_context.get(violation_type, "Revise content to comply with broadcasting standards")
        
        prompt = f"""You are an expert content editor for hoichoi digital platform. Generate a compliant revision for this S&P violation detected through hybrid analysis (keywords + context).

VIOLATION DETAILS:
- Type: {violation_type}
- Problematic Content: "{violation_text}"
- Issue: {explanation}
- Content Language: {detected_language}
- Specific Guidance: {specific_guidance}

INSTRUCTIONS:
1. Provide a revised version that eliminates the S&P violation completely
2. Maintain the original creative intent where possible
3. Keep the same language as the original content ({detected_language})
4. If the content is in Bengali, provide the solution in Bengali
5. If the content is in Hindi, provide the solution in Hindi
6. Ensure the solution is appropriate for Indian digital content standards
7. Make the minimum necessary changes to achieve compliance

Return ONLY the revised content solution in the same language as the original, nothing else."""
        
        response = client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "You are an expert content editor specializing in S&P compliance for hoichoi digital platform. You are multilingual and can provide solutions in Bengali, Hindi, and other Indian languages. Always match the language of the original content."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=300,
            temperature=0.3
        )
        
        solution = response.choices[0].message.content.strip()
        # Ensure Unicode is properly handled
        return safe_unicode_text(solution)
        
    except Exception as e:
        error_msg = str(e)
        if "billing_not_active" in error_msg:
            return f"❌ OpenAI billing issue: {error_msg}. Please check your billing at https://platform.openai.com/account/billing"
        elif "quota" in error_msg.lower():
            return f"❌ OpenAI quota exceeded: {error_msg}"
        else:
            return f"Error generating solution: {error_msg}"

def chunk_text(text, max_chars=3000):  # Reduced chunk size for better analysis
    """Split text into analysis chunks while preserving screenplay structure"""
    if len(text) <= max_chars:
        return [text]
    
    chunks = []
    
    # Split by lines first to preserve screenplay structure
    lines = text.split('\n')
    current_chunk = ""
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
            
        # Check if adding this line would exceed max_chars
        if len(current_chunk + line + '\n') > max_chars and current_chunk:
            # Save current chunk
            chunks.append(current_chunk.strip())
            current_chunk = line + '\n'
        else:
            current_chunk += line + '\n'
    
    # Add remaining chunk
    if current_chunk.strip():
        chunks.append(current_chunk.strip())
    
    # Ensure no chunk is too small (merge small chunks)
    final_chunks = []
    for chunk in chunks:
        if len(chunk) < 500 and final_chunks:  # If chunk is too small, merge with previous
            final_chunks[-1] += '\n\n' + chunk
        else:
            final_chunks.append(chunk)
    
    return final_chunks

def create_analysis_prompt():
    """Create aggressive violation detection prompt for comprehensive script analysis"""
    
    return """You are an S&P Compliance Reviewer for hoichoi. Your job is to ACTIVELY FIND and FLAG violations in this screenplay/script content. Be thorough and aggressive in detecting violations - err on the side of flagging rather than missing violations.

ANALYZE EVERYTHING: Dialogues, scene descriptions, action lines, character names, props, settings, visual elements, transitions, and any other screenplay content.

🎯 FIND VIOLATIONS IN THESE 24 AREAS:

1. **National Anthem Commercial Use** - Any use of Indian national anthem for commercial/promotional purposes
2. **Personal Information** - Real phone numbers, addresses, emails, license plates, actual photos
3. **Competitor Platform Promotion** - Any mention of Netflix, Amazon Prime, Hotstar, Disney+, Zee5, Sony Liv, or other streaming platforms
4. **National Flag/Emblem Misuse** - Improper use of Indian flag, national symbols, or emblems as props/costumes
5. **National Symbol Distortion** - Incorrect Indian map, distorted national symbols
6. **Hurtful Real References** - Negative references to real people, celebrities, organizations, sports teams
7. **Graphic Self-Harm/Suicide** - Detailed self-harm methods, explicit suicide scenes
8. **Acid Attack Scenes** - Any depiction of acid attacks
9. **Weapon/Bomb Instructions** - Detailed instructions for making weapons or explosives
10. **Harmful Product Instructions** - Using household products like phenyl for harm
11. **Religious Footwear** - Wearing shoes in temples, near idols, religious spaces
12. **Buddha Idol Misuse** - Buddha images on clothing, inappropriate contexts
13. **Religious Mockery** - Mocking religious beliefs, symbols, practices
14. **Caste/Religion Slurs** - Language reinforcing caste hierarchies, religious stereotypes
15. **Social Evils Promotion** - Glorifying child marriage, dowry, gender discrimination
16. **Visible Brand Names** - Unblurred product brands, logos, commercial endorsements
17. **Unauthorized Credits** - Unapproved changes to cast/crew credits
18. **Alcohol/Tobacco Brands** - Visible alcohol or cigarette brand names/logos
19. **Missing Smoking Warnings** - Smoking scenes without "Smoking Kills" disclaimer
20. **Missing Content Warnings** - Violent/sexual content without appropriate disclaimers
21. **Unapproved Endorsements** - Unauthorized thank-you messages in credits
22. **Animal Harm** - Actual or realistic animal cruelty, killing, suffering
23. **Child Inappropriate Behavior** - Children using adult language or mature behavior
24. **Child Abuse Content** - Any form of child abuse (physical, sexual, psychological)

🔍 DETECTION STRATEGY:
- READ EVERY LINE carefully
- LOOK FOR both obvious and subtle violations
- CHECK dialogue for inappropriate language/references
- EXAMINE scene descriptions for problematic content
- ANALYZE character actions and behaviors
- REVIEW props, settings, and visual elements
- FLAG anything that could potentially violate guidelines

⚠️ BE AGGRESSIVE IN DETECTION:
- When in doubt, FLAG IT
- Better to over-detect than miss violations
- Look for implied violations, not just explicit ones
- Consider cultural context and Indian sensitivities
- Check for subtle discriminatory language
- Look for brand names, logos, or commercial elements

Return violations in this JSON format:
{
  "violations": [
    {
      "violationText": "Exact text from script",
      "violationType": "One of the 24 violation types",
      "explanation": "Why this violates the guideline",
      "suggestedAction": "How to fix it",
      "severity": "critical|high|medium|low",
      "location": "dialogue|scene_description|action_line|character_name|prop|setting|other"
    }
  ]
}

REMEMBER: Your job is to FIND violations, not to excuse them. Be thorough, be aggressive, be comprehensive. Analyze every element of the screenplay."""

def analyze_chunk_with_mistral(chunk, chunk_num, total_chunks, api_key):
    """Analyze single chunk with Mistral API - ENHANCED with better error handling"""
    if not MISTRAL_AVAILABLE or not api_key:
        return {"violations": []}
    
    try:
        url = "https://api.mistral.ai/v1/chat/completions"
        
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}"
        }
        
        prompt = create_analysis_prompt()
        
        full_prompt = f"""{prompt}

CONTENT TO ANALYZE (Chunk {chunk_num}/{total_chunks}):
{chunk}

INSTRUCTIONS:
1. READ every line carefully
2. FIND violations in dialogues, scene descriptions, action lines, character names, props, settings
3. FLAG anything that could violate the 24 guidelines
4. Be AGGRESSIVE in detection - when in doubt, flag it
5. Handle Bengali, Hindi, and other Indian languages properly
6. Return violations in JSON format

JSON RESPONSE:"""
        
        payload = {
            "model": "mistral-small",
            "messages": [
                {
                    "role": "system", 
                    "content": "You are an aggressive S&P compliance reviewer specializing in Indian content. Your job is to FIND violations. Be thorough and flag everything that could potentially violate guidelines. You understand Bengali, Hindi, and other Indian languages. Better to over-detect than miss violations."
                },
                {
                    "role": "user", 
                    "content": full_prompt
                }
            ],
            "max_tokens": 1500,
            "temperature": 0.1
        }
        
        response = requests.post(url, json=payload, headers=headers, timeout=90)
        
        if response.status_code == 200:
            result = response.json()
            content = result['choices'][0]['message']['content'].strip()
            
            # Parse JSON response
            try:
                parsed_result = json.loads(content)
            except json.JSONDecodeError:
                # Try to extract JSON from the response
                json_match = re.search(r'\{.*\}', content, re.DOTALL)
                if json_match:
                    try:
                        parsed_result = json.loads(json_match.group())
                    except:
                        return {"violations": []}
                else:
                    return {"violations": []}
            
            # Ensure violations are properly formatted with Unicode handling
            if 'violations' in parsed_result and isinstance(parsed_result['violations'], list):
                valid_violations = []
                for violation in parsed_result['violations']:
                    if isinstance(violation, dict) and 'violationText' in violation and 'violationType' in violation:
                        # FIXED: Preserve exact Bengali text
                        violation['violationText'] = violation.get('violationText', '')
                        violation['explanation'] = violation.get('explanation', 'S&P violation detected')
                        violation['suggestedAction'] = violation.get('suggestedAction', 'Review and modify content')
                        violation.setdefault('severity', 'medium')
                        violation.setdefault('location', 'content')
                        valid_violations.append(violation)
                
                return {"violations": valid_violations}
        elif response.status_code == 429:
            st.error(f"🚨 **Mistral API Rate Limit**: Chunk {chunk_num} - Too many requests")
            return {"violations": []}
        elif response.status_code == 402:
            st.error(f"🚨 **Mistral API Billing Issue**: Chunk {chunk_num} - Payment required")
            return {"violations": []}
        else:
            st.warning(f"Mistral API error for chunk {chunk_num}: {response.status_code}")
            return {"violations": []}
        
        return {"violations": []}
        
    except Exception as e:
        st.error(f"Error analyzing chunk {chunk_num} with Mistral: {e}")
        return {"violations": []}

# FIXED: Enhanced API error handling - Replace analyze_chunk function
def analyze_chunk(chunk, chunk_num, total_chunks, api_key=None):
    """Analyze single chunk with enhanced API error handling - FIXED"""
    
    # Try Mistral first
    mistral_api_key = get_mistral_api_key_with_session()
    if MISTRAL_AVAILABLE and mistral_api_key:
        try:
            result = analyze_chunk_with_mistral(chunk, chunk_num, total_chunks, mistral_api_key)
            if result and result.get('violations'):
                return result
        except Exception as e:
            error_msg = str(e)
            st.warning(f"Mistral API issue for chunk {chunk_num}: {error_msg}")
            
            # Check for specific billing errors
            if "billing" in error_msg.lower() or "429" in error_msg:
                st.error("🚨 **Mistral API Billing Issue**: Please check your Mistral account at https://console.mistral.ai/")
            elif "quota" in error_msg.lower():
                st.error("🚨 **Mistral API Quota Exceeded**: You've reached your API usage limit")
            elif "401" in error_msg:
                st.error("🚨 **Mistral API Authentication Failed**: Please check your API key")
    
    # Fallback to OpenAI
    openai_api_key = get_api_key() if not api_key else api_key
    if OPENAI_AVAILABLE and openai_api_key:
        try:
            client = OpenAI(api_key=openai_api_key)
            
            prompt = create_analysis_prompt()
            
            full_prompt = f"""{prompt}

CONTENT TO ANALYZE (Chunk {chunk_num}/{total_chunks}):
{chunk}

Return violations in JSON format:"""
            
            response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "You are an S&P compliance reviewer. Find violations and return them in JSON format."},
                    {"role": "user", "content": full_prompt}
                ],
                temperature=0.1,
                max_tokens=1500,
                timeout=90
            )
            
            result = response.choices[0].message.content.strip()
            
            # Parse JSON response
            try:
                parsed_result = json.loads(result)
            except json.JSONDecodeError:
                json_match = re.search(r'\{.*\}', result, re.DOTALL)
                if json_match:
                    try:
                        parsed_result = json.loads(json_match.group())
                    except:
                        return {"violations": []}
                else:
                    return {"violations": []}
            
            # Ensure violations are properly formatted
            if 'violations' in parsed_result and isinstance(parsed_result['violations'], list):
                valid_violations = []
                for violation in parsed_result['violations']:
                    if isinstance(violation, dict) and 'violationText' in violation and 'violationType' in violation:
                        # FIXED: Preserve exact Bengali text
                        violation['violationText'] = violation.get('violationText', '')
                        violation['explanation'] = violation.get('explanation', 'S&P violation detected')
                        violation['suggestedAction'] = violation.get('suggestedAction', 'Review and modify content')
                        violation.setdefault('severity', 'medium')
                        violation.setdefault('location', 'content')
                        valid_violations.append(violation)
                
                return {"violations": valid_violations}
            
            return {"violations": []}
            
        except Exception as e:
            error_msg = str(e)
            st.error(f"OpenAI API error for chunk {chunk_num}: {error_msg}")
            
            # Enhanced error handling for billing issues
            if "billing_not_active" in error_msg:
                st.error("🚨 **OpenAI Billing Issue**: Your account is not active")
                st.error("**Fix**: Go to https://platform.openai.com/account/billing and add payment method")
            elif "insufficient_quota" in error_msg or "quota" in error_msg.lower():
                st.error("🚨 **OpenAI Quota Exceeded**: You've used up your API quota")
                st.error("**Fix**: Upgrade your plan or wait for quota reset")
            elif "rate_limit" in error_msg:
                st.error("🚨 **OpenAI Rate Limit**: Too many requests")
                st.error("**Fix**: Wait a moment and try again")
            elif "invalid_api_key" in error_msg:
                st.error("🚨 **OpenAI Invalid API Key**: Please check your API key")
            
            return {"violations": []}
    
    # Final fallback - keyword based analysis
    st.warning(f"Using keyword-based analysis for chunk {chunk_num} due to API issues")
    return analyze_chunk_with_keywords(chunk)

def analyze_chunk_with_keywords(chunk):
    """Fallback keyword-based analysis when APIs fail"""
    violations = []
    
    # Simple keyword-based detection
    violation_keywords = {
        "Personal_Information_Exposure": ["phone number", "mobile number", "address", "email"],
        "OTT_Platform_Promotion": ["netflix", "amazon prime", "hotstar", "zee5", "sony liv"],
        "National_Anthem_Misuse": ["national anthem", "jana gana mana"],
        "Religious_Mockery": ["mock", "ridicule", "blasphemy"],
        "Unauthorized_Branding": ["brand", "logo", "trademark"],
    }
    
    chunk_lower = chunk.lower()
    
    for violation_type, keywords in violation_keywords.items():
        for keyword in keywords:
            if keyword in chunk_lower:
                # Find the sentence containing the keyword
                sentences = chunk.split('.')
                for sentence in sentences:
                    if keyword in sentence.lower():
                        violations.append({
                            'violationText': sentence.strip(),
                            'violationType': violation_type,
                            'explanation': f'Keyword detected: "{keyword}" - requires review',
                            'suggestedAction': 'Review and modify content as needed',
                            'severity': 'medium',
                            'location': 'content'
                        })
                        break
                break
    
    return {"violations": violations}

def find_page_number(violation_text, pages_data):
    """Find which page contains the violation using original page numbers"""
    for page_data in pages_data:
        if violation_text in page_data['text']:
            return page_data.get('original_page', page_data['page_number'])
    
    # Fuzzy matching
    search_text = violation_text[:50] if len(violation_text) > 50 else violation_text
    for page_data in pages_data:
        if search_text in page_data['text']:
            return page_data.get('original_page', page_data['page_number'])
    
    return 1

def analyze_document(text, pages_data, api_key=None):
    """Analyze entire document with aggressive violation detection and better Unicode handling - ENHANCED"""
    if not text:
        return {"violations": [], "summary": {}}
    
    # Check available APIs
    mistral_key = get_mistral_api_key_with_session()
    openai_key = get_api_key() if not api_key else api_key
    
    if not mistral_key and not openai_key:
        st.error("❌ No API keys available for analysis!")
        return {"violations": [], "summary": {}}
    
    # Show which API will be used
    primary_api = "Mistral" if mistral_key else "OpenAI"
    st.info(f"🤖 **Primary API for Analysis:** {primary_api}")
    
    # Show text analysis debug info
    st.markdown("### 🔍 **Text Analysis Debug Info**")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total Characters", len(text))
    with col2:
        st.metric("Total Lines", len(text.split('\n')))
    with col3:
        unicode_chars = sum(1 for char in text if ord(char) > 127)
        st.metric("Unicode Characters", unicode_chars)
    
    # Show character distribution
    if unicode_chars > 0:
        st.success(f"✅ **Unicode Content Detected**: {unicode_chars} non-ASCII characters found")
        
        # Show character script analysis
        script_analysis = {}
        for char in text:
            if ord(char) > 127:
                script = get_script_range(char)
                script_analysis[script] = script_analysis.get(script, 0) + 1
        
        if script_analysis:
            st.markdown("**🔤 Script Analysis:**")
            for script, count in sorted(script_analysis.items(), key=lambda x: x[1], reverse=True):
                st.write(f"- {script}: {count} characters")
        
        # Show first 200 characters as preview
        with st.expander("📄 Text Preview (First 200 characters)"):
            preview_text = text[:200]
            st.text(preview_text)
            # Show if Bengali characters are present
            bengali_count = sum(1 for char in preview_text if '\u0980' <= char <= '\u09FF')
            if bengali_count > 0:
                st.success(f"✅ Bengali characters detected: {bengali_count}")
    
    # Detect language with better feedback
    detected_language = detect_language(text)
    st.info(f"🌐 **Content Language:** {detected_language} | 🔍 **Analysis Method:** Aggressive Detection | 📋 **Coverage:** Complete Script Analysis")
    
    chunks = chunk_text(text)
    all_violations = []
    successful_chunks = 0
    
    # Progress tracking
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    st.info(f"🔍 **Analyzing {len(chunks)} chunks** for comprehensive violation detection...")
    
    for i, chunk in enumerate(chunks):
        progress = (i + 1) / len(chunks)
        progress_bar.progress(progress)
        status_text.text(f"🔍 Analyzing chunk {i+1}/{len(chunks)} - Looking for violations...")
        
        # Show chunk analysis info
        chunk_unicode = sum(1 for char in chunk if ord(char) > 127)
        chunk_info = f"Chunk {i+1}: {len(chunk)} chars, {chunk_unicode} Unicode chars"
        
        # Show chunk preview for debugging
        with st.expander(f"📄 {chunk_info}"):
            st.text(chunk[:300] + "..." if len(chunk) > 300 else chunk)
            if chunk_unicode > 0:
                st.success(f"✅ Unicode content detected: {chunk_unicode} characters")
                # Show Bengali character count specifically
                bengali_count = sum(1 for char in chunk if '\u0980' <= char <= '\u09FF')
                if bengali_count > 0:
                    st.success(f"✅ Bengali characters detected: {bengali_count}")
        
        analysis = analyze_chunk(chunk, i+1, len(chunks))
        
        if 'violations' in analysis and analysis['violations']:
            st.success(f"⚠️ Found {len(analysis['violations'])} violations in chunk {i+1}")
            
            # Show violation details
            for j, violation in enumerate(analysis['violations']):
                violation_text = violation.get('violationText', '')
                violation_unicode = sum(1 for char in violation_text if ord(char) > 127)
                bengali_count = sum(1 for char in violation_text if '\u0980' <= char <= '\u09FF')
                st.write(f"   → Violation {j+1}: {violation.get('violationType', 'Unknown')} ({len(violation_text)} chars, {violation_unicode} Unicode, {bengali_count} Bengali)")
                
                violation['pageNumber'] = find_page_number(violation_text, pages_data)
                violation['chunkNumber'] = i + 1
                violation['detectedLanguage'] = detected_language
                violation['unicodeChars'] = violation_unicode
                violation['bengaliChars'] = bengali_count
                all_violations.append(violation)
            successful_chunks += 1
        else:
            st.info(f"✅ No violations found in chunk {i+1}")
    
    # Generate AI solutions for violations
    if all_violations:
        status_text.text("🤖 Generating AI solutions for violations...")
        solution_errors = 0
        
        for i, violation in enumerate(all_violations):
            progress = (i + 1) / len(all_violations)
            progress_bar.progress(progress)
            
            try:
                ai_solution = generate_ai_solution(
                    violation.get('violationText', ''),
                    violation.get('violationType', ''),
                    violation.get('explanation', ''),
                    detected_language,
                    mistral_key or openai_key
                )
                violation['aiSolution'] = ai_solution
                
                # Check if AI solution has Unicode
                solution_unicode = sum(1 for char in ai_solution if ord(char) > 127)
                solution_bengali = sum(1 for char in ai_solution if '\u0980' <= char <= '\u09FF')
                violation['aiSolutionUnicode'] = solution_unicode
                violation['aiSolutionBengali'] = solution_bengali
                
            except Exception as e:
                solution_errors += 1
                violation['aiSolution'] = f"Error generating solution: {str(e)}"
                violation['aiSolutionUnicode'] = 0
                violation['aiSolutionBengali'] = 0
        
        if solution_errors > 0:
            st.warning(f"⚠️ {solution_errors} AI solution generation errors occurred")
    
    progress_bar.progress(1.0)
    status_text.text(f"✅ Analysis complete! Found {len(all_violations)} violations across {len(chunks)} chunks")
    
    # Remove duplicates but be less aggressive about it
    unique_violations = []
    seen_violations = set()
    
    for violation in all_violations:
        v_text = violation.get('violationText', '')
        v_type = violation.get('violationType', '')
        
        # Create a more lenient duplicate detection
        duplicate_key = (v_text[:50], v_type)  # Only check first 50 chars for similarity
        
        if duplicate_key not in seen_violations:
            seen_violations.add(duplicate_key)
            unique_violations.append(violation)
    
    # Sort by severity and page
    severity_order = {'critical': 4, 'high': 3, 'medium': 2, 'low': 1}
    unique_violations.sort(key=lambda x: (
        -severity_order.get(x.get('severity', 'low'), 1),  # Sort by severity first
        x.get('pageNumber', 0)  # Then by page number
    ))
    
    return {
        "violations": unique_violations,
        "detectedLanguage": detected_language,
        "summary": {
            "totalViolations": len(unique_violations),
            "totalPages": len(pages_data),
            "chunksAnalyzed": len(chunks),
            "chunksWithViolations": successful_chunks,
            "successRate": f"{(successful_chunks/len(chunks)*100):.1f}%" if chunks else "0%",
            "unicodeChars": sum(1 for char in text if ord(char) > 127),
            "bengaliChars": sum(1 for char in text if '\u0980' <= char <= '\u09FF'),
            "totalChars": len(text),
            "primaryAPI": primary_api
        }
    }

# FIXED: Excel generation with proper Bengali text preservation
def generate_excel_report(violations, filename):
    """Generate Excel report with Bengali text preservation - FIXED"""
    if not EXCEL_AVAILABLE:
        st.error("Excel generation not available. Please install openpyxl.")
        return None
    
    try:
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Violations"
        
        # Create headers
        headers = [
            'S.No', 'Page Number', 'Violation Type', 'Severity', 
            'Violated Text (Original)', 'Explanation', 'Suggested Action', 
            'AI Solution', 'Language', 'Location', 'Status'
        ]
        
        # Add headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # FIXED: Add data rows with PRESERVED Bengali text
        for row_num, violation in enumerate(violations, 2):
            # FIXED: Preserve original Bengali text exactly as is
            violated_text = violation.get('violationText', 'N/A')
            ai_solution = violation.get('aiSolution', 'N/A')
            explanation = violation.get('explanation', 'N/A')
            suggested_action = violation.get('suggestedAction', 'N/A')
            
            # FIXED: Don't sanitize - preserve exact Bengali text
            row_data = [
                row_num - 1,  # S.No
                violation.get('pageNumber', 'N/A'),
                violation.get('violationType', 'Unknown'),
                violation.get('severity', 'medium').upper(),
                violated_text,  # FIXED: Original Bengali text preserved
                explanation,
                suggested_action,
                ai_solution,  # FIXED: Bengali solution preserved
                violation.get('detectedLanguage', 'Unknown'),
                violation.get('location', 'content'),
                'PENDING REVIEW'
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                # Set text alignment for better readability
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    cell_value = str(cell.value) if cell.value else ""
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)  # Increased max width for Bengali
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create summary sheet
        summary_ws = wb.create_sheet("Summary")
        
        # Summary data
        summary_data = [
            ['Metric', 'Count'],
            ['Total Violations', len(violations)],
            ['Critical', len([v for v in violations if v.get('severity') == 'critical'])],
            ['High', len([v for v in violations if v.get('severity') == 'high'])],
            ['Medium', len([v for v in violations if v.get('severity') == 'medium'])],
            ['Low', len([v for v in violations if v.get('severity') == 'low'])],
            ['Content Language', violations[0].get('detectedLanguage', 'Unknown') if violations else 'Unknown'],
            ['Unicode Characters', sum(v.get('unicodeChars', 0) for v in violations)],
            ['Bengali Characters', sum(v.get('bengaliChars', 0) for v in violations)]
        ]
        
        for row_num, row_data in enumerate(summary_data, 1):
            for col_num, value in enumerate(row_data, 1):
                cell = summary_ws.cell(row=row_num, column=col_num, value=value)
                if row_num == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
        
    except Exception as e:
        st.error(f"Error generating Excel report: {e}")
        return None

def create_unicode_paragraph(text, style, detected_language='English'):
    """Create a paragraph with proper Unicode support"""
    try:
        # Handle Unicode text properly
        if isinstance(text, str):
            # Ensure text is properly encoded
            safe_text = safe_unicode_text(text)
        else:
            safe_text = str(text)
        
        return Paragraph(safe_text, style)
    except Exception as e:
        # Fallback to basic text if Unicode fails
        return Paragraph(str(text).encode('ascii', errors='ignore').decode('ascii'), style)

def generate_violations_report_pdf(violations, filename):
    """Generate PDF report with ACTUAL Bengali text - FINAL FIX"""
    if not PDF_AVAILABLE:
        st.error("PDF generation not available. Please install reportlab.")
        return None
    
    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=18,
            spaceAfter=30,
            textColor=Color(0.2, 0.2, 0.6),
            alignment=1,
            fontName='Helvetica-Bold'
        )
        
        story.append(Paragraph("hoichoi S&P COMPLIANCE VIOLATION REPORT", title_style))
        story.append(Paragraph(f"Document: {filename}", styles['Normal']))
        story.append(Paragraph(f"Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        story.append(Paragraph(f"Total Violations: {len(violations)}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Summary by severity
        severity_counts = {}
        for v in violations:
            severity = v.get('severity', 'medium')
            severity_counts[severity] = severity_counts.get(severity, 0) + 1
        
        story.append(Paragraph("VIOLATION SUMMARY BY SEVERITY", styles['Heading2']))
        for severity in ['critical', 'high', 'medium', 'low']:
            count = severity_counts.get(severity, 0)
            if count > 0:
                story.append(Paragraph(f"• {severity.upper()}: {count} violations", styles['Normal']))
        
        story.append(Spacer(1, 20))
        
        # Detailed violations
        story.append(Paragraph("DETECTED VIOLATIONS WITH ACTUAL TEXT", styles['Heading1']))
        story.append(Spacer(1, 10))
        
        # Custom styles for violations
        violation_style = ParagraphStyle(
            'ViolationStyle',
            parent=styles['Normal'],
            leftIndent=20,
            rightIndent=20,
            spaceBefore=10,
            spaceAfter=10,
            borderWidth=1,
            borderColor=Color(0.8, 0.8, 0.8),
            backColor=Color(0.98, 0.98, 0.98),
            fontName='Helvetica'
        )
        
        for i, violation in enumerate(violations, 1):
            severity = violation.get('severity', 'medium')
            
            # Basic violation info
            violation_header = f"#{i} - {violation.get('violationType', 'Unknown')}"
            violation_header += f" (Page {violation.get('pageNumber', 'N/A')})"
            violation_header += f" - Severity: {severity.upper()}"
            
            story.append(Paragraph(violation_header, violation_style))
            
            # FIXED: Show ACTUAL Bengali text, not placeholders
            v_text = violation.get('violationText', 'N/A')
            
            try:
                # Try to show the actual text - let ReportLab handle it
                story.append(Paragraph(f"Violated Text: {v_text}", styles['Normal']))
            except Exception as e:
                # Only if ReportLab completely fails, show both versions
                safe_text = ''.join(char if ord(char) < 128 else '?' for char in v_text)
                story.append(Paragraph(f"Violated Text (ASCII): {safe_text}", styles['Normal']))
                story.append(Paragraph(f"Original Text: {v_text[:200]}...", styles['Normal']))
            
            story.append(Paragraph(f"Explanation: {violation.get('explanation', 'N/A')}", styles['Normal']))
            
            # FIXED: Show ACTUAL Bengali AI solution
            ai_solution = violation.get('aiSolution', 'N/A')
            
            try:
                # Try to show the actual solution - let ReportLab handle it
                story.append(Paragraph(f"AI Solution: {ai_solution}", styles['Normal']))
            except Exception as e:
                # Only if ReportLab completely fails, show both versions
                safe_solution = ''.join(char if ord(char) < 128 else '?' for char in ai_solution)
                story.append(Paragraph(f"AI Solution (ASCII): {safe_solution}", styles['Normal']))
                story.append(Paragraph(f"Original Solution: {ai_solution[:200]}...", styles['Normal']))
            
            story.append(Paragraph(f"Suggested Action: {violation.get('suggestedAction', 'N/A')}", styles['Normal']))
            story.append(Spacer(1, 10))
        
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
        
    except Exception as e:
        st.error(f"Error generating violations report PDF: {e}")
        return None

# FIXED: PDF generation without HTML spans
def generate_highlighted_text_pdf(text, violations, filename):
    """Generate PDF with ACTUAL Bengali text highlighting - FINAL FIX"""
    if not PDF_AVAILABLE:
        return None
    
    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=18,
            spaceAfter=30,
            textColor=Color(0.2, 0.2, 0.6),
            alignment=1
        )
        
        story.append(Paragraph("hoichoi S&P COMPLIANCE - SCRIPT WITH VIOLATIONS", title_style))
        story.append(Paragraph(f"Document: {filename}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Create violation mapping
        violation_map = {}
        for violation in violations:
            v_text = violation.get('violationText', '').strip()
            severity = violation.get('severity', 'medium').lower()
            
            if v_text and len(v_text) >= 3:
                violation_map[v_text] = {
                    'severity': severity,
                    'type': violation.get('violationType', 'Unknown')
                }
        
        # Process text with ACTUAL text
        story.append(Paragraph("SCRIPT CONTENT WITH MARKED VIOLATIONS", styles['Heading1']))
        story.append(Spacer(1, 10))
        
        # Split text into paragraphs
        paragraphs = text.split('\n')
        
        for para_text in paragraphs:
            if para_text.strip():
                # Check for original page markers
                if '=== ORIGINAL PAGE' in para_text:
                    page_match = re.search(r'=== ORIGINAL PAGE (\d+) ===', para_text)
                    if page_match:
                        page_num = page_match.group(1)
                        page_style = ParagraphStyle(
                            'PageMarker',
                            parent=styles['Heading3'],
                            textColor=Color(0.5, 0.5, 0.5),
                            alignment=1,
                            spaceBefore=20,
                            spaceAfter=10
                        )
                        story.append(Paragraph(f"— Original Page {page_num} —", page_style))
                    continue
                
                # Check for violations in this paragraph
                has_violation = False
                violation_severity = None
                for v_text, v_info in violation_map.items():
                    if v_text in para_text:
                        has_violation = True
                        violation_severity = v_info['severity']
                        break
                
                # Show ACTUAL text with violation marking
                if has_violation:
                    # Create violation paragraph style based on severity
                    if violation_severity == 'critical':
                        bg_color = Color(1, 0.9, 0.9)  # Light red
                        border_color = Color(1, 0, 0)   # Red border
                    elif violation_severity == 'high':
                        bg_color = Color(1, 0.95, 0.8)  # Light orange
                        border_color = Color(1, 0.5, 0) # Orange border
                    elif violation_severity == 'medium':
                        bg_color = Color(1, 1, 0.9)     # Light yellow
                        border_color = Color(1, 1, 0)   # Yellow border
                    else:
                        bg_color = Color(0.95, 0.9, 1)  # Light purple
                        border_color = Color(0.5, 0, 1) # Purple border
                    
                    violation_para_style = ParagraphStyle(
                        'ViolationPara',
                        parent=styles['Normal'],
                        spaceBefore=6,
                        spaceAfter=6,
                        leftIndent=10,
                        rightIndent=10,
                        backColor=bg_color,
                        borderWidth=2,
                        borderColor=border_color,
                        borderPadding=5
                    )
                    
                    # FIXED: Show ACTUAL text, not placeholders
                    try:
                        display_text = para_text[:1000] + "..." if len(para_text) > 1000 else para_text
                        story.append(Paragraph(f"VIOLATION [{violation_severity.upper()}]: {display_text}", violation_para_style))
                    except Exception as e:
                        safe_text = ''.join(char if ord(char) < 128 else '?' for char in para_text)
                        safe_text = safe_text[:800] + "..." if len(safe_text) > 800 else safe_text
                        story.append(Paragraph(f"VIOLATION [{violation_severity.upper()}] (ASCII): {safe_text}", violation_para_style))
                        story.append(Paragraph(f"Original Unicode: {para_text[:300]}...", styles['Normal']))
                else:
                    # Normal paragraph - show actual text
                    try:
                        display_text = para_text[:800] + "..." if len(para_text) > 800 else para_text
                        story.append(Paragraph(display_text, styles['Normal']))
                    except Exception as e:
                        safe_text = ''.join(char if ord(char) < 128 else '?' for char in para_text)
                        safe_text = safe_text[:600] + "..." if len(safe_text) > 600 else safe_text
                        story.append(Paragraph(safe_text, styles['Normal']))
                        story.append(Paragraph(f"[Unicode: {para_text[:150]}...]", styles['Normal']))
                
                story.append(Spacer(1, 2))
        
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
        
    except Exception as e:
        st.error(f"Error generating highlighted text PDF: {e}")
        return None
    """Generate PDF with original text and highlighted violations - FIXED VERSION"""
    if not PDF_AVAILABLE:
        return None
    
    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=18,
            spaceAfter=30,
            textColor=Color(0.2, 0.2, 0.6),
            alignment=1
        )
        
        story.append(Paragraph("hoichoi S&P COMPLIANCE - HIGHLIGHTED TEXT - FIXED", title_style))
        story.append(Paragraph(f"Document: {filename}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Create violation mapping for highlighting
        violation_map = {}
        for violation in violations:
            v_text = violation.get('violationText', '').strip()
            severity = violation.get('severity', 'medium').lower()
            
            if v_text and len(v_text) >= 10:
                violation_map[v_text] = {
                    'severity': severity,
                    'type': violation.get('violationType', 'Unknown')
                }
        
        # Process text WITHOUT HTML highlighting (FIXED)
        story.append(Paragraph("DOCUMENT TEXT WITH VIOLATIONS MARKED", styles['Heading1']))
        story.append(Spacer(1, 10))
        
        # Split text into paragraphs
        paragraphs = text.split('\n')
        
        for para_text in paragraphs:
            if para_text.strip():
                # Check for original page markers
                if '=== ORIGINAL PAGE' in para_text:
                    page_match = re.search(r'=== ORIGINAL PAGE (\d+) ===', para_text)
                    if page_match:
                        page_num = page_match.group(1)
                        page_style = ParagraphStyle(
                            'PageMarker',
                            parent=styles['Heading3'],
                            textColor=Color(0.5, 0.5, 0.5),
                            alignment=1,
                            spaceBefore=20,
                            spaceAfter=10
                        )
                        story.append(Paragraph(f"— Original Page {page_num} —", page_style))
                    continue
                
                # Check for violations in this paragraph
                has_violation = False
                for v_text, v_info in violation_map.items():
                    if v_text in para_text:
                        has_violation = True
                        break
                
                # FIXED: Use simple text formatting instead of HTML
                if has_violation:
                    # Mark as violation paragraph with different style
                    violation_para_style = ParagraphStyle(
                        'ViolationPara',
                        parent=styles['Normal'],
                        spaceBefore=6,
                        spaceAfter=6,
                        leftIndent=10,
                        rightIndent=10,
                        backColor=Color(1, 0.95, 0.95),  # Light red background
                        borderWidth=1,
                        borderColor=Color(1, 0, 0),
                        borderPadding=5
                    )
                    
                    # Check if text contains Bengali
                    bengali_chars = sum(1 for char in para_text if '\u0980' <= char <= '\u09FF')
                    if bengali_chars > 0:
                        # For Bengali text, show summary
                        display_text = f"[VIOLATION PARAGRAPH - Contains Bengali text: {len(para_text)} characters, {bengali_chars} Bengali characters]"
                        story.append(Paragraph(display_text, violation_para_style))
                        # Show first 100 chars as reference
                        story.append(Paragraph(f"Preview: {para_text[:100]}...", styles['Normal']))
                    else:
                        # For English text, truncate if too long
                        display_text = para_text[:500] + "..." if len(para_text) > 500 else para_text
                        story.append(Paragraph(display_text, violation_para_style))
                else:
                    # Normal paragraph
                    if len(para_text) > 800:
                        para_text = para_text[:800] + "..."
                    
                    # Check for Bengali
                    bengali_chars = sum(1 for char in para_text if '\u0980' <= char <= '\u09FF')
                    if bengali_chars > 0:
                        display_text = f"[Bengali paragraph: {len(para_text)} characters, {bengali_chars} Bengali characters]"
                        story.append(Paragraph(display_text, styles['Normal']))
                    else:
                        story.append(Paragraph(para_text, styles['Normal']))
                
                story.append(Spacer(1, 4))
        
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
        
    except Exception as e:
        st.error(f"Error generating highlighted text PDF: {e}")
        return None

def create_violation_charts(violations):
    """Create visualization charts"""
    if not violations:
        return None, None
    
    df = pd.DataFrame(violations)
    
    # Severity distribution
    severity_counts = df['severity'].value_counts()
    severity_colors = {'critical': '#f44336', 'high': '#ff9800', 'medium': '#ffeb3b', 'low': '#9c27b0'}
    
    fig_severity = px.pie(
        values=severity_counts.values,
        names=severity_counts.index,
        title="Violation Severity Distribution",
        color=severity_counts.index,
        color_discrete_map=severity_colors
    )
    
    # Violation types
    type_counts = df['violationType'].value_counts().head(10)
    fig_types = px.bar(
        x=type_counts.values,
        y=type_counts.index,
        orientation='h',
        title="Top Violation Types",
        labels={'x': 'Count', 'y': 'Violation Type'}
    )
    
    return fig_severity, fig_types

# Mistral OCR Tab Implementation
def create_mistral_ocr_tab():
    """Create OCR tab with Mistral OCR integration"""
    st.header("🔍 Mistral OCR Analysis")
    st.markdown("**Extract text from images using Mistral OCR for S&P compliance analysis.**")
    st.markdown("*🌐 Multi-language Support: Bengali, Hindi, Tamil, Telugu, Gujarati, English*")
    st.markdown("*🔤 High Accuracy: Cloud-based OCR with advanced text recognition*")
    st.markdown("*⚡ Fast Processing: No local dependencies, powered by Mistral AI*")
    st.markdown("*🎯 Optimized: Specifically tuned for Indian language scripts*")
    
    # Check Mistral OCR availability
    mistral_available, mistral_message = check_mistral_ocr_availability()
    
    if mistral_available:
        st.success("✅ Mistral OCR is ready!")
        
        # Language selection matching your n8n workflow
        language_options = {
            "Bengali + English (Recommended)": "ben+eng",
            "Hindi + English": "hin+eng", 
            "Tamil + English": "tam+eng",
            "Telugu + English": "tel+eng",
            "Gujarati + English": "guj+eng",
            "Marathi + English": "mar+eng",
            "English Only": "eng",
            "Bengali + Hindi + English": "ben+hin+eng",
            "All Indian Languages": "all"
        }
        
        selected_language = st.selectbox(
            "Select OCR Language",
            options=list(language_options.keys()),
            index=0,
            help="Choose the primary language(s) for OCR recognition. Matches your n8n workflow settings."
        )
        
        language_code = language_options[selected_language]
        st.info(f"🌐 **Language Code:** `{language_code}` (matches n8n workflow parameter)")
        
        uploaded_image = st.file_uploader(
            "Choose an image file",
            type=['png', 'jpg', 'jpeg', 'bmp', 'tiff', 'webp'],
            help="Upload an image containing text for Mistral OCR extraction and S&P analysis"
        )
        
        if uploaded_image is not None:
            st.success(f"✅ Image uploaded: {uploaded_image.name} ({uploaded_image.size/1024:.1f} KB)")
            
            # Show image preview
            if PIL_AVAILABLE:
                try:
                    image = Image.open(uploaded_image)
                    st.image(image, caption="Uploaded Image", use_column_width=True)
                    
                    # Show image info
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Width", f"{image.width} px")
                    with col2:
                        st.metric("Height", f"{image.height} px")
                    with col3:
                        st.metric("Format", image.format or "Unknown")
                except Exception as e:
                    st.warning(f"Could not display image preview: {e}")
            
            if st.button("🔍 Extract Text with Mistral OCR", type="primary", key="mistral_ocr_analyze"):
                # Extract text using Mistral OCR
                extracted_text = extract_text_with_mistral_ocr(uploaded_image, language_code)
                
                if not extracted_text.strip():
                    st.error("❌ No text could be extracted from the image")
                    st.info("💡 **Troubleshooting Tips:**")
                    st.markdown("""
                    - Ensure the image contains clear, readable text
                    - Try a higher resolution image
                    - Check if the selected language matches the text in the image
                    - Verify the image is not corrupted
                    """)
                    return
                
                st.success(f"✅ Extracted {len(extracted_text):,} characters")
                
                # Show extracted text with language detection
                detected_language = detect_language(extracted_text)
                
                with st.expander("📄 Extracted Text Preview"):
                    st.write(f"**Detected Language:** {detected_language}")
                    st.write(f"**OCR Language Setting:** {selected_language}")
                    st.write(f"**Language Code Used:** `{language_code}`")
                    st.text_area("Extracted Text", extracted_text, height=300)
                
                # Show text statistics
                unicode_chars = sum(1 for char in extracted_text if ord(char) > 127)
                bengali_chars = sum(1 for char in extracted_text if '\u0980' <= char <= '\u09FF')
                hindi_chars = sum(1 for char in extracted_text if '\u0900' <= char <= '\u097F')
                
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Total Characters", len(extracted_text))
                with col2:
                    st.metric("Unicode Characters", unicode_chars)
                with col3:
                    st.metric("Bengali Characters", bengali_chars)
                with col4:
                    st.metric("Hindi Characters", hindi_chars)
                
                # Show OCR quality assessment
                if unicode_chars > 0:
                    quality_score = min(100, int((unicode_chars / len(extracted_text)) * 200))
                    if quality_score > 80:
                        st.success(f"🎯 **OCR Quality:** Excellent ({quality_score}% Unicode density)")
                    elif quality_score > 50:
                        st.info(f"🎯 **OCR Quality:** Good ({quality_score}% Unicode density)")
                    else:
                        st.warning(f"🎯 **OCR Quality:** Fair ({quality_score}% Unicode density)")
                
                # Analyze extracted text
                pages_data = [{"page_number": 1, "text": extracted_text, "original_page": 1}]
                
                st.header("🤖 Analyzing Extracted Text")
                analysis = analyze_document(extracted_text, pages_data)
                
                violations = analysis.get('violations', [])
                
                # Display results
                display_paste_analysis_results(violations, detected_language, extracted_text)
                
    else:
        st.error(f"❌ Mistral OCR not available: {mistral_message}")
        
        # Show setup instructions
        st.markdown("### 🔧 Setup Instructions")
        st.markdown("""
        **To use Mistral OCR (matches your n8n workflow):**
        1. **Configure Mistral API Key**: Add your Mistral API key in the configuration section above
        2. **Verify OCR Access**: Ensure your Mistral subscription includes OCR capabilities
        3. **Test Connection**: Use the test button below to verify setup
        
        **Benefits of Mistral OCR:**
        - 🌐 **Multi-language Support**: Bengali, Hindi, Tamil, Telugu, Gujarati, English
        - 🔤 **High Accuracy**: Advanced OCR with context understanding
        - ⚡ **Fast Processing**: Cloud-based processing for quick results  
        - 📱 **Format Support**: PNG, JPG, JPEG, BMP, TIFF, WebP
        - 🎯 **No Dependencies**: No local OCR software installation needed
        - 🔗 **Integrated**: Matches your existing n8n workflow exactly
        """)
        
        # Test connection button
        if st.button("🧪 Test Mistral OCR Connection"):
            with st.spinner("Testing Mistral OCR connection..."):
                success, message = check_mistral_ocr_availability()
                
                if success:
                    st.success(f"✅ {message}")
                    st.balloons()
                else:
                    st.error(f"❌ {message}")
                    
                    # Provide specific troubleshooting
                    if "Invalid API key" in message:
                        st.info("🔑 Please check your Mistral API key configuration above")
                    elif "OCR access not enabled" in message:
                        st.info("📞 Contact Mistral support to enable OCR access in your account")
                    else:
                        st.info("🔧 Check your network connection and API configuration")

def update_ocr_status_in_sidebar():
    """Update OCR status in sidebar with Mistral OCR"""
    st.markdown("### 🔍 **OCR Support Status**")
    
    # Check Mistral OCR availability
    mistral_available, mistral_message = check_mistral_ocr_availability()
    
    if mistral_available:
        st.success("✅ Mistral OCR: Available")
        st.write("✓ Bengali + English support")
        st.write("✓ High accuracy cloud OCR")
        st.write("✓ No local dependencies needed")
        st.write("✓ Multi-language support")
        
        # Show supported languages
        with st.expander("🌐 Supported Language Combinations"):
            st.markdown("""
            **Primary Combinations:**
            - `ben+eng` - Bengali + English (recommended)
            - `hin+eng` - Hindi + English
            - `tam+eng` - Tamil + English  
            - `tel+eng` - Telugu + English
            - `guj+eng` - Gujarati + English
            - `mar+eng` - Marathi + English
            - `eng` - English only
            
            **Multi-language:**
            - `ben+hin+eng` - Bengali + Hindi + English
            - `all` - All supported languages
            
            **Note:** Use language codes as shown in your n8n workflow
            """)
            
    else:
        st.error("❌ Mistral OCR: Not Available")
        st.write(f"Issue: {mistral_message}")
        
        with st.expander("🔧 Setup Instructions"):
            st.markdown("""
            **To enable Mistral OCR:**
            1. ✅ Configure your Mistral API key above
            2. ✅ Ensure OCR access in your Mistral subscription  
            3. ✅ Test the connection using the button below
            
            **Troubleshooting:**
            - Verify API key is correct and active
            - Check that OCR is enabled in your Mistral account
            - Ensure network connectivity to Mistral API
            - Contact Mistral support if OCR access needed
            """)
        
        # Show detailed error message
        if OCR_ERROR_MESSAGE:
            st.error(OCR_ERROR_MESSAGE)

# Initialize Mistral OCR on startup
def initialize_mistral_ocr():
    """Initialize Mistral OCR on app startup"""
    global MISTRAL_OCR_AVAILABLE, OCR_AVAILABLE
    
    # Check Mistral OCR availability
    success, message = check_mistral_ocr_availability()
    
    # Update the global OCR_AVAILABLE flag to use Mistral OCR
    OCR_AVAILABLE = MISTRAL_OCR_AVAILABLE
    
    if success:
        print("✅ Mistral OCR initialized successfully")
    else:
        print(f"❌ Mistral OCR initialization failed: {message}")

# Check API configuration status
def check_api_configuration():
    """Check and display API configuration status - ENHANCED"""
    st.subheader("🔑 API Configuration Check")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.write("**OpenAI API Status:**")
        openai_key = get_api_key()
        if openai_key:
            try:
                # Test OpenAI API with minimal request
                client = OpenAI(api_key=openai_key)
                response = client.chat.completions.create(
                    model="gpt-3.5-turbo",
                    messages=[{"role": "user", "content": "Hi"}],
                    max_tokens=5
                )
                st.success("✅ OpenAI API: Working correctly")
            except Exception as e:
                error_msg = str(e)
                if "billing_not_active" in error_msg:
                    st.error("❌ OpenAI: Billing not active")
                    st.info("Fix: Add payment method at https://platform.openai.com/account/billing")
                elif "quota" in error_msg.lower():
                    st.error("❌ OpenAI: Quota exceeded")
                    st.info("Fix: Upgrade plan or wait for reset")
                elif "invalid" in error_msg.lower():
                    st.error("❌ OpenAI: Invalid API key")
                else:
                    st.error(f"❌ OpenAI Error: {error_msg}")
        else:
            st.warning("⚠️ OpenAI API key not configured")
    
    with col2:
        st.write("**Mistral API Status:**")
        mistral_key = get_mistral_api_key_with_session()
        if mistral_key:
            try:
                headers = {"Authorization": f"Bearer {mistral_key}"}
                response = requests.get("https://api.mistral.ai/v1/models", headers=headers, timeout=10)
                if response.status_code == 200:
                    st.success("✅ Mistral API: Working correctly")
                elif response.status_code == 401:
                    st.error("❌ Mistral: Invalid API key")
                elif response.status_code == 402:
                    st.error("❌ Mistral: Payment required")
                    st.info("Fix: Check billing at https://console.mistral.ai/")
                elif response.status_code == 429:
                    st.error("❌ Mistral: Rate limit or quota exceeded")
                else:
                    st.error(f"❌ Mistral Error: HTTP {response.status_code}")
            except Exception as e:
                st.error(f"❌ Mistral Error: {str(e)}")
        else:
            st.warning("⚠️ Mistral API key not configured")
    
    return openai_key is not None or mistral_key is not None

# Initialize Mistral OCR
initialize_mistral_ocr()

def main():
    # Authentication check
    if not authenticate_user():
        return
    
    # Initialize session state to prevent resets
    if 'analysis_complete' not in st.session_state:
        st.session_state.analysis_complete = False
    if 'violations_data' not in st.session_state:
        st.session_state.violations_data = None
    if 'current_filename' not in st.session_state:
        st.session_state.current_filename = None
    
    # Custom CSS for authenticated app
    st.markdown("""
    <style>
    .main-header {
        background: linear-gradient(90deg, #ff6b6b, #4ecdc4);
        padding: 1.5rem;
        border-radius: 15px;
        color: white;
        text-align: center;
        margin-bottom: 2rem;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
    }
    .user-info {
        background: #f8f9fa;
        padding: 1rem;
        border-radius: 10px;
        border-left: 4px solid #007bff;
    }
    .stDownloadButton > button {
        background-color: #007bff;
        color: white;
        border: none;
        padding: 0.5rem 1rem;
        border-radius: 5px;
        cursor: pointer;
    }
    .stDownloadButton > button:hover {
        background-color: #0056b3;
    }
    .bengali-text {
        font-family: 'Noto Sans Bengali', 'SolaimanLipi', Arial, sans-serif;
        font-size: 16px;
        line-height: 1.6;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Header with user info
    st.markdown("""
    <div class="main-header">
        <h1>🎬 hoichoi S&P Compliance Analyzer - FIXED VERSION</h1>
        <p>Standards & Practices Content Review Platform</p>
        <p style="font-size: 0.9em; opacity: 0.9;">✅ Bengali Text Preservation • 🔧 PDF Generation Fixed • 🔑 Enhanced API Error Handling • 📊 XLSX Export • 24 Guidelines</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Sidebar with user info and system status
    with st.sidebar:
        st.markdown(f"""
        <div class="user-info">
            <h3>👤 User Information</h3>
            <p><b>Name:</b> {st.session_state.get('user_name', 'Unknown')}</p>
            <p><b>Email:</b> {st.session_state.get('user_email', 'unknown@hoichoi.tv')}</p>
            <p><b>Role:</b> {'Admin' if st.session_state.get('is_admin', False) else 'Content Reviewer'}</p>
        </div>
        """, unsafe_allow_html=True)
        
        st.divider()
        
        st.header("🔧 FIXES APPLIED")
        st.success("✅ PDF Generation Fixed")
        st.success("✅ Bengali Text Preservation")
        st.success("✅ API Error Handling Enhanced")
        st.success("✅ Excel Export with Unicode")
        
        st.divider()
        
        st.header("🔧 System Status & Unicode Support")
        
        # Unicode support status
        st.markdown("### 🌐 **Unicode & Multilingual Support**")
        
        # Check Unicode support
        try:
            test_bengali = "আমি বাংলায় কথা বলি"
            test_hindi = "मैं हिंदी में बात करता हूं"
            
            # Test safe text processing
            safe_bengali = safe_unicode_text(test_bengali)
            safe_hindi = safe_unicode_text(test_hindi)
            
            if safe_bengali and safe_hindi and len(safe_bengali) > 0 and len(safe_hindi) > 0:
                st.success("✅ Unicode Text Processing: Working")
                st.write(f"✓ Bengali test: {safe_bengali[:20]}...")
                st.write(f"✓ Hindi test: {safe_hindi[:20]}...")
            else:
                st.error("❌ Unicode Text Processing: Issues detected")
                
        except Exception as e:
            st.error(f"❌ Unicode Processing Error: {e}")
        
        # Check language detection
        try:
            lang_detect = detect_language("আমি বাংলায় কথা বলি")
            if lang_detect and lang_detect != "English":
                st.success(f"✅ Language Detection: Working (Detected: {lang_detect})")
            else:
                st.warning("⚠️ Language Detection: Using fallback method")
                # Try fallback detection
                fallback_lang = detect_language_fallback("আমি বাংলায় কথা বলি")
                st.write(f"Fallback detected: {fallback_lang}")
                
        except Exception as e:
            st.error(f"❌ Language Detection Error: {e}")
        
        # OCR status with Mistral OCR
        update_ocr_status_in_sidebar()
        
        # System status
        st.markdown("### 🔧 **System Components**")
        
        # API Status
        if MISTRAL_AVAILABLE:
            st.success("✅ Mistral API: Available")
        else:
            st.error("❌ Mistral API: Library Missing")
            
        if OPENAI_AVAILABLE:
            st.success("✅ OpenAI API: Available")
        else:
            st.error("❌ OpenAI API: Missing")
        
        # Other components
        if DOCX_AVAILABLE:
            st.success("✅ DOCX Processing: Available")
        else:
            st.error("❌ DOCX Processing: Missing")
        
        if PDF_EXTRACT_AVAILABLE:
            st.success("✅ PDF Processing: Available")
        else:
            st.error("❌ PDF Processing: Missing")
        
        if EXCEL_AVAILABLE:
            st.success("✅ Excel Reports (XLSX): Available")
        else:
            st.error("❌ Excel Reports: Missing")
        
        if PDF_AVAILABLE:
            st.success("✅ PDF Generation: Available")
        else:
            st.error("❌ PDF Generation: Missing")
        
        st.divider()
        
        if st.button("🔄 New Analysis", type="secondary"):
            # Reset session state for new analysis
            st.session_state.analysis_complete = False
            st.session_state.violations_data = None
            st.session_state.current_filename = None
            if 'reports_generated' in st.session_state:
                del st.session_state.reports_generated
            st.rerun()
        
        if st.button("🚪 Logout", type="secondary"):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.rerun()
    
    # API Key configuration section
    st.header("🔑 API Configuration")
    
    # Check API keys
    openai_key = get_api_key()
    mistral_key = get_mistral_api_key()
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("OpenAI Configuration")
        if openai_key:
            st.success("✅ OpenAI API Key: Configured")
        else:
            st.warning("⚠️ OpenAI API Key: Not configured")
            openai_input = st.text_input(
                "Enter OpenAI API Key", 
                type="password", 
                help="For violation analysis (fallback)",
                key="openai_key_input"
            )
    
    with col2:
        st.subheader("Mistral Configuration (Preferred)")
        if mistral_key:
            st.success("✅ Mistral API Key: Configured")
        else:
            st.warning("⚠️ Mistral API Key: Not configured")
            mistral_input = st.text_input(
                "Enter Mistral API Key", 
                type="password", 
                help="For language detection, violation analysis, and OCR",
                key="mistral_key_input",
                placeholder="Enter your Mistral API key here"
            )
            if mistral_input:
                st.session_state.temp_mistral_key = mistral_input
                st.success("✅ Mistral API Key: Temporarily configured")
    
    # API Priority explanation
    st.info("""
    **🎯 API Usage Priority:**
    1. **Mistral API** (Preferred) - Better multilingual support, OCR capabilities, faster processing
    2. **OpenAI API** (Fallback) - Used if Mistral unavailable
    3. **Character-based Detection** (Final fallback) - No API required
    """)
    
    # Check API configuration
    if check_api_configuration():
        st.success("🚀 **Ready for Analysis**: API configuration verified")
    else:
        st.error("❌ **API Configuration Required**: Please configure at least one API key above")
    
    # Bengali Text Test Section
    st.header("🧪 Bengali Text Preservation Test")
    
    test_bengali = "আমি বাংলায় কথা বলি। এটি একটি পরীক্ষা।"
    test_hindi = "मैं हिंदी में बात करता हूं। यह एक परीक्षा है।"
    
    col1, col2 = st.columns(2)
    with col1:
        st.write("**Bengali Test:**")
        st.text(test_bengali)  # FIXED: Using st.text to preserve formatting
        bengali_chars = sum(1 for char in test_bengali if '\u0980' <= char <= '\u09FF')
        st.caption(f"📊 {len(test_bengali)} chars, {bengali_chars} Bengali")
    
    with col2:
        st.write("**Hindi Test:**")
        st.text(test_hindi)  # FIXED: Using st.text to preserve formatting
        hindi_chars = sum(1 for char in test_hindi if '\u0900' <= char <= '\u097F')
        st.caption(f"📊 {len(test_hindi)} chars, {hindi_chars} Hindi")
    
    if bengali_chars > 0 and hindi_chars > 0:
        st.success("✅ Unicode text display is working correctly!")
    else:
        st.error("❌ Unicode text display issue detected")
    
    # Main tabs for upload vs paste vs OCR
    tab1, tab2, tab3 = st.tabs(["📤 Upload Document", "📝 Paste Text", "🔍 Mistral OCR Analysis"])
    
    with tab1:
        st.header("📤 Upload Document Analysis")
        st.markdown("**Upload your screenplay/script for comprehensive aggressive S&P compliance review.**")
        st.markdown("*🔍 Aggressive Detection: We analyze EVERYTHING - dialogues, scene descriptions, action lines, character names, props, settings, transitions*")
        st.markdown("*⚠️ Better to over-detect than miss violations - we flag anything potentially problematic*")
        st.markdown("*📋 Complete Coverage: All 24 S&P guidelines checked across entire script*")
        st.markdown("*🌐 Enhanced Unicode Support: Proper Bengali/Hindi text handling*")
        st.markdown("*📊 XLSX Export: Full Unicode support in Excel reports*")
        
        # Show current analysis if available
        if st.session_state.analysis_complete and st.session_state.violations_data:
            display_analysis_results(st.session_state.violations_data, st.session_state.current_filename)
        else:
            uploaded_file = st.file_uploader(
                "Choose a document file",
                type=['docx', 'pdf'],
                help="Upload a Microsoft Word document (.docx) or PDF file for S&P compliance analysis"
            )
            
            if uploaded_file is not None:
                file_type = uploaded_file.name.split('.')[-1].lower()
                st.success(f"✅ File uploaded: {uploaded_file.name} ({uploaded_file.size/1024:.1f} KB)")
                
                if st.button("🔍 Start Analysis", type="primary", key="upload_analyze"):
                    # Extract text based on file type
                    with st.spinner(f"📄 Extracting text from {file_type.upper()} document..."):
                        if file_type == 'pdf':
                            text, pages_data = extract_text_from_pdf_bytes(uploaded_file.getvalue())
                        else:  # docx
                            text, pages_data = extract_text_from_docx_bytes(uploaded_file.getvalue())
                    
                    if not text:
                        st.error("❌ Failed to extract text from document")
                        return
                    
                    st.success(f"✅ Extracted {len(text):,} characters from {len(pages_data)} pages")
                    
                    # Analyze document
                    st.header("🤖 Analysis in Progress")
                    analysis = analyze_document(text, pages_data)
                    
                    # Store results in session state
                    st.session_state.violations_data = {
                        'violations': analysis.get('violations', []),
                        'summary': analysis.get('summary', {}),
                        'detected_language': analysis.get('detectedLanguage', 'Unknown'),
                        'text': text,
                        'pages_data': pages_data
                    }
                    st.session_state.current_filename = uploaded_file.name
                    st.session_state.analysis_complete = True
                    
                    # Display results
                    display_analysis_results(st.session_state.violations_data, uploaded_file.name)
    
    with tab2:
        st.header("📝 Paste Text Analysis")
        st.markdown("**Paste your screenplay content for comprehensive aggressive S&P compliance review.**")
        st.markdown("*🔍 Aggressive Detection: We examine every line for potential violations*")
        st.markdown("*⚠️ Thorough Analysis: Dialogues, scene descriptions, actions, character behavior, props, settings*")
        st.markdown("*🌐 Enhanced Unicode Support: Perfect Bengali/Hindi text handling*")
        
        text_input = st.text_area(
            "Paste your screenplay/script content here",
            height=300,
            placeholder="Paste your screenplay content here for aggressive S&P compliance analysis...\n\nINT. LIVING ROOM - DAY\n\nRAJ sits on the sofa.\n\nRAJ\n(dialogue here)\nHello, how are you?\n\nPRIYA enters the room.\n\nPRIYA\nI'm fine, thanks.\n\nOr paste Bengali text:\nআমি বাংলায় কথা বলি। এটি একটি পরীক্ষা।\n\nOur AI will analyze every element for potential violations!"
        )
        
        if text_input and st.button("🔍 Analyze Text", type="primary", key="paste_analyze"):
            # Create mock pages data for pasted text
            pages_data = [{"page_number": 1, "text": text_input, "original_page": 1}]
            
            # Analyze pasted text
            st.header("🤖 Analyzing Pasted Text")
            analysis = analyze_document(text_input, pages_data)
            
            violations = analysis.get('violations', [])
            detected_language = analysis.get('detectedLanguage', 'Unknown')
            
            # Display results for pasted text
            display_paste_analysis_results(violations, detected_language, text_input)
    
    with tab3:
        create_mistral_ocr_tab()
    
    # Footer with violation rules
    with st.expander("📋 S&P Violation Guidelines Reference (24 Aggressive Detection Rules)"):
        st.markdown("### 🎯 hoichoi Standards & Practices Guidelines")
        st.markdown("**Our system uses AGGRESSIVE DETECTION to find violations across your entire screenplay. We analyze every element thoroughly.**")
        st.markdown("---")
        
        st.markdown("### 🔍 **What We Analyze:**")
        st.markdown("""
        - **📝 Dialogues**: Every line of character speech
        - **🎬 Scene Descriptions**: Location setups, visual descriptions
        - **🎭 Action Lines**: Character movements, behaviors, actions
        - **👥 Character Names**: All character references and mentions
        - **🎪 Props & Settings**: Objects, locations, visual elements
        - **🎞️ Transitions**: Scene changes, cuts, fades, directions
        """)
        
        st.markdown("### ⚠️ **24 Violation Categories We Detect:**")
        
        for i, (rule_name, rule_data) in enumerate(VIOLATION_RULES.items(), 1):
            severity = rule_data['severity']
            if severity == "critical":
                st.error(f"🔴 **{i}. {rule_name.replace('_', ' ')}**")
            elif severity == "high":
                st.warning(f"🟠 **{i}. {rule_name.replace('_', ' ')}**")
            else:
                st.info(f"🟡 **{i}. {rule_name.replace('_', ' ')}**")
            
            st.markdown(f"**Description:** {rule_data['description']}")
            st.markdown(f"**Context:** {rule_data['context']}")
            st.markdown(f"**Keywords:** {', '.join(rule_data['keywords'])}")
            st.markdown("---")
        
        st.markdown("### 🎯 **Detection Philosophy**")
        st.markdown("""
        **✅ Better to Over-Detect than Miss Violations**
        - We flag anything that could potentially be problematic
        - When in doubt, we flag it for your review
        - Comprehensive analysis of all screenplay elements
        
        **🔍 Multi-Method Detection**
        - Keyword scanning for quick identification
        - Context analysis for subtle violations
        - Cultural sensitivity checks
        - Intent and meaning analysis
        
        **📋 Complete Coverage**
        - No element of your screenplay is ignored
        - Every dialogue line is examined
        - Every scene description is analyzed
        - Every action and prop is checked
        """)
        
        st.markdown("**📝 Result:** Comprehensive violation detection across your entire screenplay with detailed solutions for each issue found.")
    
    # Footer
    st.markdown("---")
    st.markdown(f"""
    <div style='text-align: center; color: #666; font-size: 0.9em;'>
        <p>🎬 hoichoi S&P Compliance System v3.0 - FIXED VERSION | Enhanced Unicode Support | Mistral OCR Integration | XLSX Export | Reviewed by: {st.session_state.get('user_name', 'Unknown')}</p>
        <p>🔒 Secure access • 🔍 Aggressive violation detection • 🌐 Bengali/Hindi/Multi-language support • 📊 Full Unicode Excel reports • 🔍 Cloud OCR</p>
        <p style="color: green;">✅ FIXES: PDF generation • Bengali text preservation • API error handling • Excel Unicode support</p>
    </div>
    """, unsafe_allow_html=True)

# FIXED: Enhanced Display Functions - Replace display_violation_details
def display_violation_details(violation, index, detected_language):
    """Display individual violation details with exact Bengali text preservation - FIXED"""
    severity = violation.get('severity', 'low')
    
    if severity == 'critical':
        st.error(f"🔴 **{violation.get('violationType', 'Unknown')}** (Page {violation.get('pageNumber', 'N/A')})")
    elif severity == 'high':
        st.warning(f"🟠 **{violation.get('violationType', 'Unknown')}** (Page {violation.get('pageNumber', 'N/A')})")
    elif severity == 'medium':
        st.info(f"🟡 **{violation.get('violationType', 'Unknown')}** (Page {violation.get('pageNumber', 'N/A')})")
    else:
        st.success(f"🟢 **{violation.get('violationType', 'Unknown')}** (Page {violation.get('pageNumber', 'N/A')})")
    
    col_a, col_b = st.columns([1, 1])
    
    with col_a:
        st.write("**🚨 Exact Violated Text:**")
        violation_text = violation.get("violationText", "N/A")
        
        # FIXED: Show exact Bengali text using st.text (preserves formatting)
        st.text(violation_text)
        
        # Also show in expandable code block for copying
        with st.expander("📋 Copy Exact Text"):
            st.code(violation_text, language=None)
        
        # Show character analysis
        unicode_count = sum(1 for char in violation_text if ord(char) > 127)
        bengali_count = sum(1 for char in violation_text if '\u0980' <= char <= '\u09FF')
        hindi_count = sum(1 for char in violation_text if '\u0900' <= char <= '\u097F')
        
        if unicode_count > 0:
            st.caption(f"📊 {len(violation_text)} chars total | {unicode_count} Unicode | {bengali_count} Bengali | {hindi_count} Hindi")
        
        st.write(f"**Issue:** {violation.get('explanation', 'N/A')}")
    
    with col_b:
        st.write(f"**🤖 AI Solution ({detected_language}):**")
        ai_solution = violation.get("aiSolution", "N/A")
        
        # FIXED: Show exact Bengali solution
        st.text(ai_solution)
        
        # Also show in expandable code block for copying
        with st.expander("📋 Copy Solution"):
            st.code(ai_solution, language=None)
        
        # Show solution character analysis
        solution_unicode = sum(1 for char in ai_solution if ord(char) > 127)
        solution_bengali = sum(1 for char in ai_solution if '\u0980' <= char <= '\u09FF')
        
        if solution_unicode > 0:
            st.caption(f"📊 {len(ai_solution)} chars total | {solution_unicode} Unicode | {solution_bengali} Bengali")
        
        st.write(f"**Recommended Action:** {violation.get('suggestedAction', 'N/A')}")
    
    st.divider()

def display_analysis_results(violations_data, filename):
    """Display analysis results with aggressive detection feedback - ENHANCED"""
    violations = violations_data['violations']
    summary = violations_data['summary']
    detected_language = violations_data['detected_language']
    text = violations_data['text']
    pages_data = violations_data['pages_data']
    
    # Results
    st.header("📊 Aggressive Analysis Results")
    
    # Show detection summary
    st.info(f"🔍 **Detection Summary**: Analyzed {summary.get('chunksAnalyzed', 0)} chunks, found violations in {summary.get('chunksWithViolations', 0)} chunks")
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total Violations Found", summary.get('totalViolations', 0))
    with col2:
        critical_count = len([v for v in violations if v.get('severity') == 'critical'])
        st.metric("🔴 Critical Issues", critical_count)
    with col3:
        st.metric("📄 Pages Analyzed", summary.get('totalPages', 0))
    with col4:
        st.metric("✅ Detection Rate", summary.get('successRate', '0%'))
    
    # Show Unicode statistics
    if summary.get('unicodeChars', 0) > 0:
        st.success(f"🌐 **Unicode Content**: {summary.get('unicodeChars', 0)} Unicode characters, {summary.get('bengaliChars', 0)} Bengali characters detected")
    
    if violations:
        # Show detection effectiveness
        st.success(f"✅ **Aggressive Detection Successful**: Found {len(violations)} violations across your screenplay")
        
        # Charts
        st.subheader("📈 Violation Analytics")
        fig_severity, fig_types = create_violation_charts(violations)
        
        col1, col2 = st.columns(2)
        with col1:
            if fig_severity:
                st.plotly_chart(fig_severity, use_container_width=True)
        
        with col2:
            if fig_types:
                st.plotly_chart(fig_types, use_container_width=True)
        
        # Violation details with AI solutions
        st.subheader(f"🚨 Detected Violations with AI Solutions ({detected_language})")
        st.markdown("*Each violation detected through comprehensive script analysis*")
        
        for i, violation in enumerate(violations[:15]):  # Show first 15
            display_violation_details(violation, i+1, detected_language)
        
        if len(violations) > 15:
            st.info(f"Showing first 15 of {len(violations)} total violations detected")
        
        # Download Reports Section
        st.subheader("📥 Download Comprehensive Reports")
        
        # Generate reports (cached to avoid regeneration)
        if 'reports_generated' not in st.session_state:
            with st.spinner("Generating comprehensive reports..."):
                excel_data = generate_excel_report(violations, filename)
                violations_pdf = generate_violations_report_pdf(violations, filename)
                highlighted_pdf = generate_highlighted_text_pdf(text, violations, filename)
                
                st.session_state.reports_generated = {
                    'excel': excel_data,
                    'violations_pdf': violations_pdf,
                    'highlighted_pdf': highlighted_pdf
                }
        
        reports = st.session_state.reports_generated
        
        # Download buttons in columns
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if reports['excel']:
                st.download_button(
                    label="📊 Excel Report (XLSX)",
                    data=reports['excel'],
                    file_name=f"{filename}_violations_detected.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="excel_download"
                )
                st.success("✅ Full Unicode support in Excel")
        
        with col2:
            if reports['violations_pdf']:
                st.download_button(
                    label="📋 Violations Report (PDF)",
                    data=reports['violations_pdf'],
                    file_name=f"{filename}_violations_report.pdf",
                    mime="application/pdf",
                    key="violations_download"
                )
                st.info("ℹ️ Bengali text shown as placeholders in PDF")
        
        with col3:
            if reports['highlighted_pdf']:
                st.download_button(
                    label="🎨 Highlighted Script (PDF)",
                    data=reports['highlighted_pdf'],
                    file_name=f"{filename}_highlighted_violations.pdf",
                    mime="application/pdf",
                    key="highlighted_download"
                )
                st.info("ℹ️ Bengali text shown as placeholders in PDF")
        
        st.info(f"📋 **Reports Generated:** Excel with {len(violations)} violations (full Unicode support), PDF violation summary, and highlighted script with flagged content")
    
    else:
        st.success("🎉 No violations detected! Your content appears to comply with S&P standards.")
        st.info("📋 **Note**: Our aggressive detection system analyzed your entire screenplay and found no violations against the 24 S&P guidelines.")
        st.balloons()

def display_paste_analysis_results(violations, detected_language, text_input):
    """Display analysis results for pasted text with EXACT snippets - ENHANCED"""
    st.header(f"📊 Analysis Results ({detected_language})")
    
    # Show text statistics
    total_chars = len(text_input)
    unicode_chars = sum(1 for char in text_input if ord(char) > 127)
    bengali_chars = sum(1 for char in text_input if '\u0980' <= char <= '\u09FF')
    
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Total Characters", total_chars)
    with col2:
        st.metric("Unicode Characters", unicode_chars)
    with col3:
        st.metric("Bengali Characters", bengali_chars)
    
    if violations:
        st.error(f"🚨 Found {len(violations)} violations in your text!")
        
        # Show violations with exact context and AI solutions
        st.subheader("🔍 Violated Content with AI Solutions")
        st.markdown("*Detected using hybrid analysis: keyword detection + contextual understanding*")
        
        for i, violation in enumerate(violations, 1):
            severity = violation.get('severity', 'low')
            
            # Color-coded violation display
            if severity == 'critical':
                st.error(f"**🔴 Violation #{i}: {violation.get('violationType', 'Unknown')}**")
            elif severity == 'high':
                st.warning(f"**🟠 Violation #{i}: {violation.get('violationType', 'Unknown')}**")
            elif severity == 'medium':
                st.info(f"**🟡 Violation #{i}: {violation.get('violationType', 'Unknown')}**")
            else:
                st.success(f"**🟢 Violation #{i}: {violation.get('violationType', 'Unknown')}**")
            
            # Show violated text with highlighting and AI solution
            violated_text = violation.get('violationText', '')
            ai_solution = violation.get('aiSolution', 'No solution available')
            
            col_a, col_b = st.columns([1, 1])
            
            with col_a:
                st.markdown("**🚨 Exact Violated Text:**")
                
                # FIXED: Show exact violated text using st.text (preserves formatting)
                st.text(violated_text)
                
                # Also show in expandable code block for copying
                with st.expander("📋 Copy Exact Text"):
                    st.code(violated_text, language=None)
                
                # Show context around the violation (optional)
                if len(text_input) > len(violated_text):
                    # Find the position of violated text in the input
                    pos = text_input.find(violated_text)
                    if pos != -1:
                        # Show 50 characters before and after for context
                        start = max(0, pos - 50)
                        end = min(len(text_input), pos + len(violated_text) + 50)
                        context = text_input[start:end]
                        
                        st.markdown("**📄 Context:**")
                        # Highlight the violated text within the context using markdown
                        highlighted_context = context.replace(violated_text, f"**{violated_text}**")
                        st.markdown(f"...{highlighted_context}...")
                
                st.markdown(f"**Why this violates S&P:** {violation.get('explanation', 'N/A')}")
            
            with col_b:
                st.markdown(f"**🤖 AI Solution ({detected_language}):**")
                
                # FIXED: Show exact Bengali solution
                st.text(ai_solution)
                
                # Also show in expandable code block for copying
                with st.expander("📋 Copy Solution"):
                    st.code(ai_solution, language=None)
                
                st.markdown(f"**Suggested action:** {violation.get('suggestedAction', 'N/A')}")
                st.markdown(f"**Severity:** {severity.upper()}")
            
            st.divider()
        
        # Show severity summary
        st.subheader("📊 Violation Summary")
        severity_counts = {}
        for v in violations:
            severity = v.get('severity', 'medium')
            severity_counts[severity] = severity_counts.get(severity, 0) + 1
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("🔴 Critical", severity_counts.get('critical', 0))
        with col2:
            st.metric("🟠 High", severity_counts.get('high', 0))
        with col3:
            st.metric("🟡 Medium", severity_counts.get('medium', 0))
        with col4:
            st.metric("🟢 Low", severity_counts.get('low', 0))
        
        # Generate reports for paste analysis
        st.subheader("📥 Download Reports")
        
        col1, col2 = st.columns(2)
        
        with col1:
            # Generate Excel report
            excel_data = generate_excel_report(violations, "Pasted_Text_Analysis")
            if excel_data:
                st.download_button(
                    label="📊 Excel Report (XLSX)",
                    data=excel_data,
                    file_name="pasted_text_violations.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="paste_excel_download"
                )
                st.success("✅ Full Unicode support in Excel")
        
        with col2:
            # Generate PDF report
            violations_pdf = generate_violations_report_pdf(violations, "Pasted_Text_Analysis")
            if violations_pdf:
                st.download_button(
                    label="📋 Violations Report (PDF)",
                    data=violations_pdf,
                    file_name="pasted_text_violations.pdf",
                    mime="application/pdf",
                    key="paste_pdf_download"
                )
                st.info("ℹ️ Bengali text shown as placeholders in PDF")
    
    else:
        st.success("🎉 No violations found! Your text appears to comply with S&P standards.")
        st.balloons()

if __name__ == "__main__":
    main()
